from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.exceptions import NotFoundError
from backend.repositories.cliente_repository import ClienteRepository
from backend.repositories.proposta_repository import PropostaRepository
from backend.repositories.proposta_item_repository import PropostaItemRepository
from backend.repositories.proposta_item_composicao_repository import (
    PropostaItemComposicaoRepository,
)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


class PropostaExportService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db
        self.proposta_repo = PropostaRepository(db)
        self.cliente_repo = ClienteRepository(db)
        self.item_repo = PropostaItemRepository(db)
        self.composicao_repo = PropostaItemComposicaoRepository(db)

    async def gerar_excel(self, proposta_id: UUID) -> bytes:
        proposta = await self.proposta_repo.get_by_id(proposta_id)
        if proposta is None:
            raise NotFoundError("Proposta", str(proposta_id))

        cliente = await self.cliente_repo.get_by_id(proposta.cliente_id)
        itens = await self.item_repo.list_by_proposta(proposta_id)

        wb = Workbook()
        # Aba 1: Capa
        capa = wb.active
        capa.title = "Capa"
        capa["A1"] = "Codigo"
        capa["B1"] = "Cliente"
        capa["A2"] = proposta.codigo
        capa["B2"] = proposta.codigo
        capa["A3"] = "Titulo"
        capa["B3"] = proposta.titulo or ""
        capa["A4"] = "Status"
        capa["B4"] = proposta.status.value if hasattr(proposta.status, "value") else str(proposta.status)
        capa["A5"] = "Cliente"
        capa["B5"] = cliente.nome_fantasia if cliente else ""
        if cliente and getattr(cliente, "cnpj", None):
            capa["A6"] = "CNPJ"
            capa["B6"] = cliente.cnpj
        capa["A8"] = "Total Direto"
        capa["B8"] = float(proposta.total_direto or 0)
        capa["A9"] = "Total Indireto"
        capa["B9"] = float(proposta.total_indireto or 0)
        capa["A10"] = "Total Geral"
        capa["B10"] = float(proposta.total_geral or 0)
        for col in ("A", "B"):
            capa.column_dimensions[col].width = 28

        # Aba 2: Quadro-Resumo (agregado por TipoRecurso)
        resumo = wb.create_sheet("Quadro-Resumo")
        _write_header(resumo, 1, ["Tipo de Recurso", "Custo Total"])
        agregado: dict[str, Decimal] = {}
        composicoes_por_item: dict = {}
        for item in itens:
            comps = await self.composicao_repo.list_by_proposta_item(item.id)
            composicoes_por_item[item.id] = comps
            for c in comps:
                tipo = c.tipo_recurso.value if c.tipo_recurso else "OUTRO"
                agregado[tipo] = agregado.get(tipo, Decimal("0")) + (c.custo_total_insumo or Decimal("0"))
        for row, (tipo, valor) in enumerate(sorted(agregado.items()), start=2):
            resumo.cell(row=row, column=1, value=tipo)
            resumo.cell(row=row, column=2, value=float(valor))
        resumo.column_dimensions["A"].width = 24
        resumo.column_dimensions["B"].width = 18

        # Aba 3: CPU
        cpu = wb.create_sheet("CPU")
        _write_header(cpu, 1, ["Codigo", "Descricao", "Unidade", "Qtd", "Custo Direto", "Custo Indireto", "Preco Unitario", "Preco Total"])
        for row, item in enumerate(itens, start=2):
            cpu.cell(row=row, column=1, value=item.codigo)
            cpu.cell(row=row, column=2, value=item.descricao)
            cpu.cell(row=row, column=3, value=item.unidade_medida)
            cpu.cell(row=row, column=4, value=float(item.quantidade or 0))
            cpu.cell(row=row, column=5, value=float(item.custo_direto_unitario or 0))
            cpu.cell(row=row, column=6, value=float(item.custo_indireto_unitario or 0))
            cpu.cell(row=row, column=7, value=float(item.preco_unitario or 0))
            cpu.cell(row=row, column=8, value=float(item.preco_total or 0))

        # Aba 4: Composicoes
        comp_ws = wb.create_sheet("Composicoes")
        _write_header(comp_ws, 1, ["Item Codigo", "Insumo", "Unidade", "Qtd Consumo", "Custo Unit", "Custo Total", "Tipo Recurso", "Nivel"])
        row = 2
        for item in itens:
            for c in composicoes_por_item.get(item.id, []):
                comp_ws.cell(row=row, column=1, value=item.codigo)
                comp_ws.cell(row=row, column=2, value=c.descricao_insumo)
                comp_ws.cell(row=row, column=3, value=c.unidade_medida)
                comp_ws.cell(row=row, column=4, value=float(c.quantidade_consumo or 0))
                comp_ws.cell(row=row, column=5, value=float(c.custo_unitario_insumo or 0))
                comp_ws.cell(row=row, column=6, value=float(c.custo_total_insumo or 0))
                comp_ws.cell(row=row, column=7, value=c.tipo_recurso.value if c.tipo_recurso else "")
                comp_ws.cell(row=row, column=8, value=c.nivel)
                row += 1

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def gerar_pdf(self, proposta_id: UUID) -> bytes:
        proposta = await self.proposta_repo.get_by_id(proposta_id)
        if proposta is None:
            raise NotFoundError("Proposta", str(proposta_id))
        cliente = await self.cliente_repo.get_by_id(proposta.cliente_id)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title=f"Proposta {proposta.codigo}")
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Proposta {proposta.codigo}", styles["Title"]),
            Spacer(1, 12),
            Paragraph(f"<b>Titulo:</b> {proposta.titulo or '-'}", styles["Normal"]),
            Paragraph(f"<b>Cliente:</b> {cliente.nome_fantasia if cliente else '-'}", styles["Normal"]),
        ]
        if cliente and getattr(cliente, "cnpj", None):
            story.append(Paragraph(f"<b>CNPJ:</b> {cliente.cnpj}", styles["Normal"]))
        story.append(Spacer(1, 18))

        totals_data = [
            ["Indicador", "Valor (R$)"],
            ["Total Direto", f"{float(proposta.total_direto or 0):,.2f}"],
            ["Total Indireto", f"{float(proposta.total_indireto or 0):,.2f}"],
            ["Total Geral", f"{float(proposta.total_geral or 0):,.2f}"],
        ]
        table = Table(totals_data, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ]))
        story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
