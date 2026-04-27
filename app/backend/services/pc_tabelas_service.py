"""ETL service: parses PC Tabelas XLSX and persists all sheets to the database."""

from __future__ import annotations

import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.pc_tabelas import (
    EtlCarga,
    PcCabecalho,
    PcEncargoItem,
    PcEquipamentoItem,
    PcEquipamentoPremissa,
    PcEpiDistribuicaoFuncao,
    PcEpiItem,
    PcFerramentaItem,
    PcMaoObraItem,
    PcMobilizacaoItem,
    PcMobilizacaoQuantidadeFuncao,
)


def _to_decimal(value: Any):
    """Safe cast to float-compatible Decimal; returns None on blank/error."""
    if value is None:
        return None
    try:
        return float(str(value).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _find_col(col_map: dict[str, int], *keywords: str) -> int | None:
    """Finds first column index matching any of the keywords (case-insensitive)."""
    for kw in keywords:
        kw_up = kw.upper()
        for hk, idx in col_map.items():
            if kw_up in hk:
                return idx
    return None


def _parse_mao_obra(ws, cabecalho_id: uuid.UUID) -> list[PcMaoObraItem]:
    """
    ABA 'MÃO DE OBRA'
    Header detection: looks for row with 'DESCRI' in it.
    """
    header_row_idx = 3  # default fallback
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if any(cell and "DESCRI" in str(cell).upper() for cell in row):
            header_row_idx = row_num
            break

    header_vals = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    col_map = {str(c).strip().upper(): i for i, c in enumerate(header_vals) if c}

    c_desc = _find_col(col_map, "DESCRI") if col_map else None
    if c_desc is None:
        c_desc = 0
    c_qtd = _find_col(col_map, "QUANT") or 1
    c_sal = _find_col(col_map, "SALARIO", "SALÁRIO") or 2
    c_reaj = _find_col(col_map, "REAJUSTE") or 3
    c_enc = _find_col(col_map, "ENCARGO") or 4
    c_peric = _find_col(col_map, "PERICULOSIDADE", "INSALUBRIDADE") or 5
    c_refei = _find_col(col_map, "REFEIÇÃO", "REFEICAO") or 6
    c_agua = _find_col(col_map, "ÁGUA", "AGUA") or 7
    c_vale = _find_col(col_map, "VALE ALIMENTAÇÃO", "VALE ALIMENTACAO") or 8
    c_saude = _find_col(col_map, "SAÚDE", "SAUDE") or 9
    c_ferr = _find_col(col_map, "FERRAMENTA") or 10
    c_seg = _find_col(col_map, "SEGURO") or 11
    c_ferias = _find_col(col_map, "FÉRIAS", "FERIAS") or 12
    c_unif = _find_col(col_map, "UNIFORME") or 13
    c_epi = _find_col(col_map, "EPI") or 14
    c_cunit = _find_col(col_map, "CUSTO UNITARIO", "UNITÁRIO") or 15
    c_cmensal = _find_col(col_map, "CUSTO MENSAL") or 16
    c_mob = _find_col(col_map, "MOBILIZAÇÃO", "MOBILIZACAO") or 18

    def _v(row: tuple, c: int | None) -> float | None:
        return _to_decimal(row[c]) if c is not None and c < len(row) else None

    items: list[PcMaoObraItem] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        desc = row[c_desc] if c_desc < len(row) else None
        if not desc:
            continue
        items.append(
            PcMaoObraItem(
                id=uuid.uuid4(),
                pc_cabecalho_id=cabecalho_id,
                descricao_funcao=str(desc).strip(),
                quantidade=_v(row, c_qtd),
                salario=_v(row, c_sal),
                previsao_reajuste=_v(row, c_reaj),
                encargos_percent=_v(row, c_enc),
                periculosidade_insalubridade=_v(row, c_peric),
                refeicao=_v(row, c_refei),
                agua_potavel=_v(row, c_agua),
                vale_alimentacao=_v(row, c_vale),
                plano_saude=_v(row, c_saude),
                ferramentas_val=_v(row, c_ferr),
                seguro_vida=_v(row, c_seg),
                abono_ferias=_v(row, c_ferias),
                uniforme_val=_v(row, c_unif),
                epi_val=_v(row, c_epi),
                custo_unitario_h=_v(row, c_cunit),
                custo_mensal=_v(row, c_cmensal),
                mobilizacao=_v(row, c_mob),
            )
        )
    return items


def _parse_equipamentos(ws, cabecalho_id: uuid.UUID):
    """
    ABA 'EQUIPAMENTOS'
    Row 2: Horas/mês (col I=8), Gasolina (col I row3), Diesel (col I row4)
    Row 6: header (CÓDIGO | EQUIPAMENTO | COMBUSTÍVEL | CONSUMO | ALUGUEL | COMBUSTÍVEL/h | MÃO DE OBRA/h | HP | HI | MÊS | | ALUGUEL MENSAL)
    Data starts row 7.
    """
    premissa = None
    items: list[PcEquipamentoItem] = []

    all_rows = list(ws.iter_rows(values_only=True))

    # Parse premissas from rows 2-4 (0-indexed: 1-3)
    horas_mes = None
    gasolina = None
    diesel = None
    for r in all_rows[:5]:
        if r and len(r) > 9:
            label = str(r[8]).lower() if r[8] else ""
            if "hora" in label:
                horas_mes = _to_decimal(r[9])
            if "gasolina" in label:
                gasolina = _to_decimal(r[9])
            if "diesel" in label:
                diesel = _to_decimal(r[9])

    premissa = PcEquipamentoPremissa(
        id=uuid.uuid4(),
        pc_cabecalho_id=cabecalho_id,
        horas_mes=horas_mes,
        preco_gasolina_l=gasolina,
        preco_diesel_l=diesel,
    )

    # Data rows start after header (row index 6+ = row 7+)
    for row in ws.iter_rows(min_row=7, values_only=True):
        equip = row[1] if len(row) > 1 else None
        if not equip:
            continue
        items.append(
            PcEquipamentoItem(
                id=uuid.uuid4(),
                pc_cabecalho_id=cabecalho_id,
                codigo=str(row[0]).strip() if row[0] else None,
                equipamento=str(equip).strip(),
                combustivel_utilizado=str(row[2]).strip() if row[2] else None,
                consumo_l_h=_to_decimal(row[3]),
                aluguel_r_h=_to_decimal(row[4]),
                combustivel_r_h=_to_decimal(row[5]),
                mao_obra_r_h=_to_decimal(row[6]),
                hora_produtiva=_to_decimal(row[7]),
                hora_improdutiva=_to_decimal(row[8]),
                mes=_to_decimal(row[9]) if len(row) > 9 else None,
                aluguel_mensal=_to_decimal(row[11]) if len(row) > 11 else None,
            )
        )
    return premissa, items


def _parse_encargos(ws, cabecalho_id: uuid.UUID, tipo: str) -> list[PcEncargoItem]:
    """
    ABA 'ENCARGOS HORISTA' or 'ENCARGOS MENSALISTA'
    Find the header row, then parse groups/items.
    """
    items: list[PcEncargoItem] = []
    current_grupo = None

    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and str(row[0]).strip().upper() not in ("GRUPOS",):
            # Could be a group label like "A", "B", "C"
            val0 = str(row[0]).strip()
            if len(val0) <= 3 and val0.isalpha():
                current_grupo = val0

        # Encargo horista: cols 1=código, 2=discriminação, 5=taxa
        # Encargo mensalista: cols 0=grupo, 1=código, 2=discriminação, 3=taxa
        if tipo == "HORISTA":
            cod = str(row[1]).strip() if row[1] else None
            disc = str(row[2]).strip() if row[2] else None
            taxa = _to_decimal(row[5]) if len(row) > 5 else None
        else:
            cod = str(row[1]).strip() if row[1] else None
            disc = str(row[2]).strip() if row[2] else None
            taxa = _to_decimal(row[3]) if len(row) > 3 else None

        if not disc or disc.lower() in ("discriminação do encargo", "grupos"):
            continue
        if disc.lower().startswith("total do grupo"):
            continue

        items.append(
            PcEncargoItem(
                id=uuid.uuid4(),
                pc_cabecalho_id=cabecalho_id,
                tipo_encargo=tipo,
                grupo=current_grupo,
                codigo_grupo=cod,
                discriminacao_encargo=disc,
                taxa_percent=taxa,
            )
        )
    return items


# EPI column mapping: which function columns exist (0-indexed from col 8)
_EPI_FUNCOES = [
    "OFICIAL", "AJUDANTE", "ELETRICISTA", "OPERADOR",
    "MÃO DE OBRA INDIRETA", "ALPINISTA", "MONTADOR ANDAIME",
]


def _parse_epi(ws, cabecalho_id: uuid.UUID):
    """
    ABA 'EPI-UNIFORME'
    Row 2 = header: (blank) | EPI | UNID | CUSTO UNITÁRIO | QTDE | VIDA ÚTIL | CUSTO COM EPI(MÊS) | (blank) | OFICIAL | AJUDANTE | ELETRICISTA | OPERADOR | MÃO DE OBRA INDIRETA | ALPINISTA | MONTADOR ANDAIME
    Data starts row 3.
    """
    epi_items: list[PcEpiItem] = []
    dist_items: list[PcEpiDistribuicaoFuncao] = []

    # Detect function columns from row 2 (index 1)
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    funcao_cols: list[tuple[int, str]] = []
    for idx, cell in enumerate(header_row):
        if cell and idx >= 8:
            funcao_cols.append((idx, str(cell).strip()))

    for row in ws.iter_rows(min_row=3, values_only=True):
        epi_name = row[1] if len(row) > 1 else None
        if not epi_name:
            continue

        epi_id = uuid.uuid4()
        epi_items.append(
            PcEpiItem(
                id=epi_id,
                pc_cabecalho_id=cabecalho_id,
                epi=str(epi_name).strip(),
                unidade=str(row[2]).strip() if row[2] else None,
                custo_unitario=_to_decimal(row[3]),
                quantidade=_to_decimal(row[4]),
                vida_util_meses=_to_decimal(row[5]),
                custo_epi_mes=_to_decimal(row[6]),
            )
        )

        for col_idx, funcao in funcao_cols:
            val = row[col_idx] if len(row) > col_idx else None
            if val is not None:
                dist_items.append(
                    PcEpiDistribuicaoFuncao(
                        id=uuid.uuid4(),
                        pc_epi_item_id=epi_id,
                        funcao=funcao,
                        aplica_flag=str(val).strip(),
                    )
                )

    return epi_items, dist_items


def _parse_ferramentas(ws, cabecalho_id: uuid.UUID) -> list[PcFerramentaItem]:
    """
    ABA 'FERRAMENTAS'
    Row 3 = header: (blank) | ITEM | DESCRIÇÃO | UNID. | QUANT. | PREÇO | PREÇO TOTAL
    Data starts row 4.
    """
    items: list[PcFerramentaItem] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        desc = row[2] if len(row) > 2 else None
        if not desc:
            continue
        items.append(
            PcFerramentaItem(
                id=uuid.uuid4(),
                pc_cabecalho_id=cabecalho_id,
                item=str(row[1]).strip() if row[1] else None,
                descricao=str(desc).strip(),
                unidade=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                quantidade=_to_decimal(row[4]) if len(row) > 4 else None,
                preco=_to_decimal(row[5]) if len(row) > 5 else None,
                preco_total=_to_decimal(row[6]) if len(row) > 6 else None,
            )
        )
    return items


def _parse_mobilizacao(ws, cabecalho_id: uuid.UUID):
    """
    ABA 'MOBILIZAÇÃO'
    Row 2 = function columns header: (blank) | (blank) | ENG | TEC | AUX IND | ADM | ENC | AJUD | OFI | OPE | ELE
    Row 5 = function names: Descrição | Função | Engenheiro | Técnico | ...
    Row 6 = Quantidade row
    Rows 8+ = exam rows with values per function column.
    """
    mob_items: list[PcMobilizacaoItem] = []
    quant_items: list[PcMobilizacaoQuantidadeFuncao] = []

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Get function header columns from row 2 (index 1)
    header_row2 = all_rows[1] if len(all_rows) > 1 else []
    funcao_cols: list[tuple[int, str]] = []
    for idx, cell in enumerate(header_row2):
        if cell and idx >= 2:
            funcao_cols.append((idx, str(cell).strip()))

    # Data rows start at row 8 (index 7)
    for row in all_rows[7:]:
        desc = row[0]
        if not desc:
            continue
        item_id = uuid.uuid4()
        valor_unitario = row[1] if len(row) > 1 else None
        mob_items.append(
            PcMobilizacaoItem(
                id=item_id,
                pc_cabecalho_id=cabecalho_id,
                descricao=str(desc).strip(),
                funcao=str(valor_unitario).strip() if valor_unitario else None,
                tipo_mao_obra=None,
            )
        )
        for col_idx, funcao in funcao_cols:
            val = row[col_idx] if len(row) > col_idx else None
            quant_items.append(
                PcMobilizacaoQuantidadeFuncao(
                    id=uuid.uuid4(),
                    pc_mobilizacao_item_id=item_id,
                    coluna_funcao=funcao,
                    quantidade=_to_decimal(val),
                )
            )

    return mob_items, quant_items


async def importar_pc_tabelas(
    db: AsyncSession,
    file_bytes: bytes,
    nome_arquivo: str,
) -> PcCabecalho:
    """
    Parse all sheets from PC Tabelas XLSX and persist to DB.
    Replaces any existing data for the same filename.
    Returns the PcCabecalho record.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    # Remove previous data with same filename
    result = await db.execute(select(PcCabecalho).where(PcCabecalho.nome_arquivo == nome_arquivo))
    existing = result.scalars().all()
    for cab in existing:
        await db.delete(cab)
    await db.flush()

    # Create ETL log
    etl = EtlCarga(
        id=uuid.uuid4(),
        fonte_arquivo=nome_arquivo,
        tipo_fonte="PC_TABELAS",
        status="EM_PROCESSAMENTO",
        iniciado_em=datetime.now(timezone.utc),
    )
    db.add(etl)
    await db.flush()

    # Create cabecalho
    cab = PcCabecalho(
        id=uuid.uuid4(),
        etl_carga_id=etl.id,
        nome_arquivo=nome_arquivo,
        versao_layout="v1",
    )
    db.add(cab)
    await db.flush()

    total_rows = 0

    # ── Mão de Obra ─────────────────────────────────────────────────────────
    for sheet_name in wb.sheetnames:
        if "MÃO DE OBRA" in sheet_name.upper() and "INDIRETA" not in sheet_name.upper():
            ws = wb[sheet_name]
            for item in _parse_mao_obra(ws, cab.id):
                db.add(item)
                total_rows += 1
            break

    # ── Equipamentos ────────────────────────────────────────────────────────
    for sheet_name in wb.sheetnames:
        if "EQUIPAMENTO" in sheet_name.upper():
            ws = wb[sheet_name]
            premissa, eq_items = _parse_equipamentos(ws, cab.id)
            db.add(premissa)
            for item in eq_items:
                db.add(item)
                total_rows += 1
            break

    # ── Encargos Horista ─────────────────────────────────────────────────────
    horista_done = False
    for sheet_name in wb.sheetnames:
        if "HORISTA" in sheet_name.upper():
            ws = wb[sheet_name]
            for item in _parse_encargos(ws, cab.id, "HORISTA"):
                db.add(item)
                total_rows += 1
            horista_done = True
            break

    # ── Encargos Mensalista ──────────────────────────────────────────────────
    mensalista_done = False
    for sheet_name in wb.sheetnames:
        if "MENSALISTA" in sheet_name.upper():
            ws = wb[sheet_name]
            for item in _parse_encargos(ws, cab.id, "MENSALISTA"):
                db.add(item)
                total_rows += 1
            mensalista_done = True
            break

    # ── ENCARGOS FALLBACK (if no HORISTA/MENSALISTA sheets found) ───────────
    if not horista_done and not mensalista_done:
        for sheet_name in wb.sheetnames:
            if "ENCARGO" in sheet_name.upper():
                ws = wb[sheet_name]
                # Try parsing as HORISTA by default if just one sheet exists
                for item in _parse_encargos(ws, cab.id, "HORISTA"):
                    db.add(item)
                    total_rows += 1
                break

    # ── EPI / Uniforme ───────────────────────────────────────────────────────
    for sheet_name in wb.sheetnames:
        if "EPI" in sheet_name.upper():
            ws = wb[sheet_name]
            epi_items, dist_items = _parse_epi(ws, cab.id)
            for item in epi_items:
                db.add(item)
                total_rows += 1
            for dist in dist_items:
                db.add(dist)
            break

    # ── Ferramentas ──────────────────────────────────────────────────────────
    for sheet_name in wb.sheetnames:
        if "FERRAMENTA" in sheet_name.upper():
            ws = wb[sheet_name]
            for item in _parse_ferramentas(ws, cab.id):
                db.add(item)
                total_rows += 1
            break

    # ── Mobilização ──────────────────────────────────────────────────────────
    for sheet_name in wb.sheetnames:
        if "MOBILIZA" in sheet_name.upper():
            ws = wb[sheet_name]
            mob_items, quant_items = _parse_mobilizacao(ws, cab.id)
            for item in mob_items:
                db.add(item)
                total_rows += 1
            for q in quant_items:
                db.add(q)
            break

    # Finalise ETL log
    etl.status = "CONCLUIDO"
    etl.finalizado_em = datetime.now(timezone.utc)
    etl.linhas_lidas = total_rows
    etl.linhas_carregadas = total_rows

    await db.commit()
    await db.refresh(cab)
    return cab

