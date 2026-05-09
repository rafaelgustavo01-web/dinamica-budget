import json
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any
from uuid import UUID

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.enums import CampoSistemaPQ
from backend.models.pq_layout import PqImportacaoMapeamento, PqLayoutCliente, PqLayoutHistorico
from backend.repositories.pq_layout_repository import PqLayoutRepository
from backend.schemas.pq_layout import PqLayoutCriarRequest

_HEADER_ALIASES = {
    CampoSistemaPQ.CODIGO: {"codigo", "código", "cod", "item", "item codigo"},
    CampoSistemaPQ.DESCRICAO: {
        "descricao", "descrição", "servico", "serviço",
        "item descricao", "item descrição",
        "descricao das atividades", "descrição das atividades",
    },
    CampoSistemaPQ.UNIDADE: {"unidade", "unid", "unid.", "und", "und.", "unidade_medida", "unidade medida"},
    CampoSistemaPQ.QUANTIDADE: {"quantidade", "qtde", "qtd", "quant", "quant.", "coeficiente", "coef", "coef."},
    CampoSistemaPQ.OBSERVACAO: {"observacao", "observação", "obs", "obs."},
}


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return " ".join(text.replace("_", " ").split())


def _infer_mapeamento(headers: list[object]) -> list[tuple[CampoSistemaPQ, str]]:
    normalized = [_normalize_header(value) for value in headers]
    encontrados: list[tuple[CampoSistemaPQ, str]] = []
    usados: set[int] = set()
    for campo, aliases in _HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            if idx in usados:
                continue
            if header in aliases:
                encontrados.append((campo, str(headers[idx]).strip()))
                usados.add(idx)
                break
    return encontrados


def _calcular_score(headers: list[object], layout: PqLayoutCliente | None) -> Decimal:
    if layout is None:
        return Decimal("0")
    normalized = [_normalize_header(value) for value in headers]
    colunas_layout = {m.coluna_planilha.strip().lower() for m in layout.mapeamentos}
    aliases: set[str] = set()
    if layout.aliases_json:
        try:
            aliases_data = json.loads(layout.aliases_json)
            for vals in aliases_data.values():
                if isinstance(vals, list):
                    aliases.update(v.strip().lower() for v in vals if isinstance(v, str))
        except json.JSONDecodeError:
            pass
    total = len(colunas_layout)
    if total == 0:
        return Decimal("0")
    encontradas = 0
    for col in colunas_layout:
        norm = _normalize_header(col)
        if norm in normalized or norm in aliases:
            encontradas += 1
    score = Decimal(encontradas) / Decimal(total)
    return Decimal(str(round(score, 4)))


class PqLayoutService:
    def __init__(self, db: AsyncSession) -> None:
        self._db = db
        self._repo = PqLayoutRepository(db)

    async def criar_ou_substituir(self, cliente_id: UUID, req: PqLayoutCriarRequest) -> PqLayoutCliente:
        await self._repo.delete_by_cliente_id(cliente_id)
        layout = PqLayoutCliente(
            id=uuid.uuid4(),
            cliente_id=cliente_id,
            nome=req.nome,
            aba_nome=req.aba_nome,
            linha_inicio=req.linha_inicio,
            is_aprovado=False,
            aliases_json=req.aliases_json,
            mapeamentos=[],
        )
        for m in req.mapeamentos:
            layout.mapeamentos.append(
                PqImportacaoMapeamento(
                    id=uuid.uuid4(),
                    campo_sistema=m.campo_sistema,
                    coluna_planilha=m.coluna_planilha,
                )
            )
        result = await self._repo.create(layout)
        await self._repo.registrar_historico(
            PqLayoutHistorico(
                id=uuid.uuid4(),
                layout_id=result.id,
                cliente_id=cliente_id,
                acao="CRIADO",
                usuario_id=None,
                detalhe_json=json.dumps({"nome": req.nome, "linha_inicio": req.linha_inicio}),
            )
        )
        return result

    async def obter_por_cliente(self, cliente_id: UUID) -> PqLayoutCliente | None:
        return await self._repo.get_by_cliente_id(cliente_id)

    async def aprovar(self, layout_id: UUID, usuario_id: UUID) -> PqLayoutCliente:
        layout = await self._repo.get_by_id(layout_id)
        if layout is None:
            from backend.core.exceptions import NotFoundError
            raise NotFoundError("PqLayoutCliente", str(layout_id))
        await self._repo.aprovar(layout, usuario_id)
        await self._repo.registrar_historico(
            PqLayoutHistorico(
                id=uuid.uuid4(),
                layout_id=layout.id,
                cliente_id=layout.cliente_id,
                acao="APROVADO",
                usuario_id=usuario_id,
                detalhe_json=None,
            )
        )
        return layout

    async def listar_historico(self, layout_id: UUID) -> list[PqLayoutHistorico]:
        return await self._repo.list_historico_by_layout(layout_id)

    def detectar_colunas_xlsx(self, filepath: str, aba_nome: str | None) -> list[str]:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[aba_nome] if aba_nome and aba_nome in wb.sheetnames else wb.active
        primeira = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        wb.close()
        return [str(c) for c in primeira if c is not None]

    def build_coluna_map(self, layout: PqLayoutCliente) -> dict[str, str]:
        return {m.campo_sistema.value: m.coluna_planilha for m in layout.mapeamentos}

    def sugerir_mapeamento(self, headers: list[object]) -> list[dict[str, Any]]:
        encontrados = _infer_mapeamento(headers)
        return [{"campo_sistema": campo.value, "coluna_planilha": coluna} for campo, coluna in encontrados]

    def calcular_score(self, headers: list[object], layout: PqLayoutCliente | None) -> Decimal:
        return _calcular_score(headers, layout)
