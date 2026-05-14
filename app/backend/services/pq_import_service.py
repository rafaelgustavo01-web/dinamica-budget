import csv
import io
import re
import uuid
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID

import openpyxl
from fastapi import UploadFile

from backend.core.exceptions import NotFoundError, ValidationError
from backend.core.logging import get_logger
from backend.models.enums import StatusImportacao, StatusMatch, StatusProposta
from backend.models.pq_layout import PqLayoutCliente
from backend.models.proposta import PqImportacao, PqItem
from backend.repositories.associacao_repository import normalize_text
from backend.repositories.pq_importacao_repository import PqImportacaoRepository
from backend.repositories.pq_item_repository import PqItemRepository
from backend.repositories.proposta_repository import PropostaRepository

logger = get_logger(__name__)

_HEADER_ALIASES = {
    "codigo": {"codigo", "código", "cod", "item", "item codigo"},
    "descricao": {
        "descricao", "descrição", "servico", "serviço",
        "item descricao", "item descrição",
        "descricao das atividades", "descrição das atividades",
    },
    "unidade": {"unidade", "unid", "unid.", "und", "und.", "unidade_medida", "unidade medida"},
    "quantidade": {"quantidade", "qtde", "qtd", "quant", "quant.", "coeficiente", "coef", "coef."},
}
_SUPPORTED_EXTENSIONS = {"csv", "xlsx"}

# Padrões de título/seção em planilhas de quantitativos (PQ)
# Inclui variações com/sem acento para robustez com planilhas de clientes diversos
_SECTION_KEYWORDS = {
    "capitulo", "capítulo", "secao", "seção", "seçao",
    "servicos", "serviços", "serviçoes",
    "titulo", "título",
    "etapa", "fase", "disciplina", "grupo", "subgrupo",
    "empreitada", "contrato", "obra", "projeto",
    "relatorio", "relatório", "relatorios", "relatórios",
    "resumo", "total", "subtotal", "geral", "sumario", "sumário",
}
_SECTION_NUMBERING_RE = re.compile(r"^\d+(\.\d+)*\s*[\.:)\-]?\s*$")
_DIGITS_ONLY_RE = re.compile(r"^[\d\s\.\,\:\-\(\)\/]+$")


def _is_likely_section_title(row: dict[str, Any]) -> bool:
    """Detecta se uma linha de planilha PQ é um título/seção (não um item de obra).

    Regras determinísticas e conservadoras — prefere manter um item legítimo
    a descartar um item válido. Um item de obra REAL sempre possui
    quantidade > 0 e unidade de medida. Títulos de seção nunca possuem.

    Heurísticas (OR):
      1. Sem quantidade válida (> 0) E sem unidade de medida.
      2. Descrição curta (<= 5 caracteres).
      3. Descrição totalmente em maiúsculas (<= 50 chars) E sem qtd E sem unidade.
      4. Descrição é apenas numeração de capítulo/seção (ex: "1.", "2.1.3") E sem qtd.
      5. Descrição começa com palavra-chave de seção (ex: "CAPÍTULO 1") E sem qtd E sem unidade.
    """
    descricao = str(row.get("descricao") or "").strip()
    if not descricao:
        return True  # linha vazia já seria pega antes, mas garante

    qtd_raw = row.get("quantidade")
    unidade_raw = row.get("unidade")

    has_qtd = False
    if qtd_raw not in (None, ""):
        try:
            qtd_val = Decimal(str(qtd_raw).strip().replace(",", "."))
            has_qtd = qtd_val > 0
        except InvalidOperation:
            has_qtd = False

    has_unidade = bool(str(unidade_raw or "").strip())

    # Regra 1: sem qtd E sem unidade → título (segura, itens reais sempre têm ambos)
    if not has_qtd and not has_unidade:
        return True

    # Regra 2: descrição muito curta E sem dados (não descarta itens curtos válidos com qtd+unidade)
    if len(descricao) <= 5 and not has_qtd and not has_unidade:
        return True

    # Se tem quantidade E unidade, é quase certamente um item real — não descarta
    if has_qtd and has_unidade:
        return False

    # Regra 3: maiúsculas curtas sem dados numéricos
    if descricao.isupper() and len(descricao) <= 50 and not has_qtd and not has_unidade:
        return True

    # Regra 4: apenas numeração de seção
    if _SECTION_NUMBERING_RE.match(descricao) and not has_qtd:
        return True

    # Regra 5: começa com palavra-chave de seção
    first_word = descricao.split()[0].lower().rstrip(".:-)")
    if first_word in _SECTION_KEYWORDS and not has_qtd and not has_unidade:
        return True

    # Regra 6: descrição composta apenas de dígitos, pontuação e espaços, sem unidade
    if _DIGITS_ONLY_RE.match(descricao) and not has_unidade:
        return True

    return False


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return " ".join(text.replace("_", " ").split())


def _decode_csv(contents: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return contents.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValidationError("Não foi possível decodificar o CSV informado.")


def _find_column_map(headers: list[object]) -> dict[str, int]:
    normalized = [_normalize_header(value) for value in headers]
    column_map: dict[str, int] = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                column_map[canonical] = idx
                break
    if "descricao" not in column_map:
        raise ValidationError("A planilha deve conter uma coluna de descrição.")
    return column_map


def _find_column_map_from_layout(headers: list[object], layout_map: dict[str, str]) -> dict[str, int]:
    normalized = [_normalize_header(value) for value in headers]
    column_map: dict[str, int] = {}
    for campo, coluna_planilha in layout_map.items():
        coluna_norm = _normalize_header(coluna_planilha)
        for idx, header in enumerate(normalized):
            if header == coluna_norm:
                column_map[campo] = idx
                break
    if "descricao" not in column_map:
        raise ValidationError("A planilha deve conter uma coluna de descrição.")
    return column_map


def _parse_decimal(value: object, default: Decimal) -> Decimal:
    if value in (None, ""):
        return default
    normalized = str(value).strip()
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation as exc:
        raise ValidationError("Quantidade inválida na planilha.", details={"valor": str(value)}) from exc


def _build_layout_map(layout: PqLayoutCliente | None) -> dict[str, str] | None:
    if layout is None:
        return None
    return {m.campo_sistema.value: m.coluna_planilha for m in layout.mapeamentos}


def _calcular_score(headers: list[object], layout: PqLayoutCliente | None) -> Decimal:
    if layout is None or not layout.mapeamentos:
        return Decimal("0")
    normalized = [_normalize_header(value) for value in headers]
    encontrados = 0
    total = len(layout.mapeamentos)
    for m in layout.mapeamentos:
        norm = _normalize_header(m.coluna_planilha)
        if norm in normalized:
            encontradas = 1
        else:
            encontradas = 0
        encontrados += encontradas
    return Decimal(str(round(encontrados / total, 4)))


class PqImportService:
    def __init__(
        self,
        proposta_repo: PropostaRepository,
        importacao_repo: PqImportacaoRepository,
        item_repo: PqItemRepository,
        pq_layout_repo=None,
    ) -> None:
        self.proposta_repo = proposta_repo
        self.importacao_repo = importacao_repo
        self.item_repo = item_repo
        self._pq_layout_repo = pq_layout_repo

    async def _resolver_layout(self, cliente_id: UUID) -> PqLayoutCliente | None:
        if self._pq_layout_repo is None:
            return None
        return await self._pq_layout_repo.get_by_cliente_id(cliente_id)

    async def preview_planilha(
        self, proposta_id: UUID, arquivo: UploadFile
    ) -> dict[str, Any]:
        """Parseia arquivo sem gravar no banco. Retorna itens sugeridos + score."""
        proposta = await self.proposta_repo.get_by_id(proposta_id)
        if not proposta:
            raise NotFoundError("Proposta", str(proposta_id))
        if not arquivo.filename:
            raise ValidationError("Arquivo inválido.")

        ext = arquivo.filename.rsplit(".", 1)[-1].lower() if "." in arquivo.filename else ""
        if ext not in _SUPPORTED_EXTENSIONS:
            raise ValidationError("Somente arquivos .csv e .xlsx são suportados.")

        contents = await arquivo.read()
        if not contents:
            raise ValidationError("Arquivo vazio.")

        layout = await self._resolver_layout(proposta.cliente_id)
        parsed_rows = self._parse_contents(contents, ext, layout=layout)

        score = Decimal("0")
        if layout is not None and parsed_rows:
            # score baseado no cabeçalho real do arquivo
            score = _calcular_score(self._last_headers, layout)

        itens: list[dict[str, Any]] = []
        linhas_ok = 0
        linhas_com_erro = 0
        linhas_ignoradas = 0
        for row in parsed_rows:
            descricao = str(row["descricao"] or "").strip()
            if not descricao:
                linhas_com_erro += 1
                itens.append({
                    "linha_planilha": row["linha_planilha"],
                    "codigo": None,
                    "descricao": "",
                    "unidade": None,
                    "quantidade": Decimal("0"),
                    "status": "ERRO",
                    "erro_msg": "Descrição vazia",
                })
                continue

            # Detecção inteligente de títulos/seções
            if _is_likely_section_title(row):
                linhas_ignoradas += 1
                itens.append({
                    "linha_planilha": row["linha_planilha"],
                    "codigo": str(row["codigo"]).strip() if row["codigo"] not in (None, "") else None,
                    "descricao": descricao,
                    "unidade": str(row["unidade"]).strip() if row["unidade"] not in (None, "") else None,
                    "quantidade": _parse_decimal(row["quantidade"], Decimal("0")),
                    "status": "IGNORADO",
                    "erro_msg": "Identificado como título/seção",
                })
                continue

            try:
                quantidade = _parse_decimal(row["quantidade"], Decimal("1"))
            except ValidationError as exc:
                linhas_com_erro += 1
                itens.append({
                    "linha_planilha": row["linha_planilha"],
                    "codigo": str(row["codigo"]).strip() if row["codigo"] not in (None, "") else None,
                    "descricao": descricao,
                    "unidade": str(row["unidade"]).strip() if row["unidade"] not in (None, "") else None,
                    "quantidade": Decimal("0"),
                    "status": "ERRO",
                    "erro_msg": str(exc),
                })
                continue

            codigo = str(row["codigo"]).strip() if row["codigo"] not in (None, "") else None
            unidade = str(row["unidade"]).strip() if row["unidade"] not in (None, "") else None
            linhas_ok += 1
            itens.append({
                "linha_planilha": row["linha_planilha"],
                "codigo": codigo,
                "descricao": descricao,
                "unidade": unidade,
                "quantidade": quantidade,
                "status": "OK",
                "erro_msg": None,
            })

        return {
            "score_confianca": score,
            "linhas_total": len(parsed_rows),
            "linhas_ok": linhas_ok,
            "linhas_com_erro": linhas_com_erro,
            "linhas_ignoradas": linhas_ignoradas,
            "itens": itens,
        }

    async def importar_planilha(self, proposta_id: uuid.UUID, arquivo: UploadFile) -> PqImportacao:
        proposta = await self.proposta_repo.get_by_id(proposta_id)
        if not proposta:
            raise NotFoundError("Proposta", str(proposta_id))
        if not arquivo.filename:
            raise ValidationError("Arquivo inválido.")

        ext = arquivo.filename.rsplit(".", 1)[-1].lower() if "." in arquivo.filename else ""
        if ext not in _SUPPORTED_EXTENSIONS:
            raise ValidationError("Somente arquivos .csv e .xlsx são suportados.")

        contents = await arquivo.read()
        if not contents:
            raise ValidationError("Arquivo vazio.")

        # Limpa itens e importações anteriores não confirmados.
        # CONFIRMADO e MANUAL são preservados (revisados pelo usuário).
        await self.item_repo.delete_nao_confirmados(proposta_id)
        await self.importacao_repo.delete_by_proposta(proposta_id)

        layout = await self._resolver_layout(proposta.cliente_id)
        layout_map = _build_layout_map(layout)
        parsed_rows = self._parse_contents(contents, ext, layout=layout)

        importacao = PqImportacao(
            id=uuid.uuid4(),
            proposta_id=proposta_id,
            nome_arquivo=arquivo.filename,
            formato=ext,
            linhas_total=len(parsed_rows),
            linhas_importadas=0,
            linhas_com_erro=0,
            status=StatusImportacao.PROCESSANDO,
        )
        importacao = await self.importacao_repo.create(importacao)

        itens: list[PqItem] = []
        linhas_com_erro = 0
        linhas_ignoradas = 0
        for row in parsed_rows:
            descricao = str(row["descricao"] or "").strip()
            if not descricao:
                linhas_com_erro += 1
                continue

            # Detecção inteligente de títulos/seções — descarta sem criar PqItem
            if _is_likely_section_title(row):
                linhas_ignoradas += 1
                logger.info(
                    "pq_import.titulo_ignorado",
                    proposta_id=str(proposta_id),
                    linha=row["linha_planilha"],
                    descricao=descricao,
                )
                continue

            try:
                quantidade = _parse_decimal(row["quantidade"], Decimal("1"))
            except ValidationError:
                linhas_com_erro += 1
                continue

            codigo = str(row["codigo"]).strip() if row["codigo"] not in (None, "") else None
            unidade = str(row["unidade"]).strip() if row["unidade"] not in (None, "") else None
            itens.append(
                PqItem(
                    proposta_id=proposta_id,
                    pq_importacao_id=importacao.id,
                    codigo_original=codigo,
                    descricao_original=descricao,
                    unidade_medida_original=unidade,
                    quantidade_original=quantidade,
                    descricao_tokens=normalize_text(descricao),
                    match_status=StatusMatch.PENDENTE,
                    linha_planilha=row["linha_planilha"],
                )
            )

        if itens:
            await self.item_repo.create_batch(itens)
            if proposta.status == StatusProposta.RASCUNHO:
                proposta.status = StatusProposta.EM_ANALISE

        importacao.linhas_importadas = len(itens)
        importacao.linhas_com_erro = linhas_com_erro
        importacao.linhas_ignoradas = linhas_ignoradas
        importacao.status = StatusImportacao.CONCLUIDO if linhas_com_erro == 0 else StatusImportacao.COM_ERROS
        await self.importacao_repo.update(importacao)
        return importacao

    def _parse_contents(
        self, contents: bytes, ext: str, layout: PqLayoutCliente | None = None
    ) -> list[dict[str, Any]]:
        if ext == "csv":
            return self._parse_csv(contents, layout=layout)
        return self._parse_xlsx(contents, layout=layout)

    def _parse_csv(self, contents: bytes, layout: PqLayoutCliente | None = None) -> list[dict[str, Any]]:
        text = _decode_csv(contents)
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return []

        layout_map = _build_layout_map(layout)
        if layout_map:
            column_map = _find_column_map_from_layout(rows[0], layout_map)
        else:
            column_map = _find_column_map(rows[0])
        self._last_headers = rows[0]

        start_line = layout.linha_inicio if layout and layout.linha_inicio else 2
        parsed_rows: list[dict[str, Any]] = []
        # CSV é 0-based internamente; linha_inicio 2 significa pular rows[0] (header) e rows[1] (linha 2 do arquivo)
        # Mas rows[0] é sempre header. Se linha_inicio > 2, pular mais linhas.
        skip = max(0, start_line - 2)
        for line_number, row in enumerate(rows[1 + skip :], start=2 + skip):
            if not any(str(cell).strip() for cell in row if cell is not None):
                continue
            parsed_rows.append(self._build_parsed_row(row, column_map, line_number))
        return parsed_rows

    def _parse_xlsx(self, contents: bytes, layout: PqLayoutCliente | None = None) -> list[dict[str, Any]]:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        worksheet = workbook[layout.aba_nome] if layout and layout.aba_nome and layout.aba_nome in workbook.sheetnames else workbook.active

        layout_map = _build_layout_map(layout)
        header_row_number = None
        header_values: list[object] = []
        linha_inicio = layout.linha_inicio if layout and layout.linha_inicio else 1

        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_number < linha_inicio:
                continue
            values = list(row)
            try:
                if layout_map:
                    _find_column_map_from_layout(values, layout_map)
                else:
                    _find_column_map(values)
                header_row_number = row_number
                header_values = values
                break
            except ValidationError:
                continue

        if header_row_number is None:
            raise ValidationError("Não foi possível identificar o cabeçalho da planilha.")

        self._last_headers = header_values

        if layout_map:
            column_map = _find_column_map_from_layout(header_values, layout_map)
        else:
            column_map = _find_column_map(header_values)
        parsed_rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            values = list(row)
            if not any(str(cell).strip() for cell in values if cell is not None):
                continue
            parsed_rows.append(self._build_parsed_row(values, column_map, row_number))
        workbook.close()
        return parsed_rows

    def _build_parsed_row(
        self,
        row: list[object],
        column_map: dict[str, int],
        line_number: int,
    ) -> dict[str, Any]:
        def _value(field: str) -> object:
            idx = column_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        return {
            "codigo": _value("codigo"),
            "descricao": _value("descricao"),
            "unidade": _value("unidade"),
            "quantidade": _value("quantidade"),
            "linha_planilha": line_number,
        }
