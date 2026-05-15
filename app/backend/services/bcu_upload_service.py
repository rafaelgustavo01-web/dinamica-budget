"""Upload individual de bases BCU por tipo, com preview e validação."""

from __future__ import annotations

import uuid
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.exceptions import NotFoundError, UnprocessableEntityError, ValidationError
from backend.core.logging import get_logger
from backend.models.base_tcpo import BaseTcpo
from backend.models.bcu import (
    BcuCabecalho,
    BcuEncargoItem,
    BcuEquipamentoItem,
    BcuEpiItem,
    BcuFerramentaItem,
    BcuMaoObraItem,
    BcuMobilizacaoItem,
)

logger = get_logger(__name__)

_VALID_TIPOS = {"mo", "equipamentos", "encargos", "exames", "epi", "ferramentas", "mobilizacao"}
_MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
_XLSX_MAGIC = b"PK\x03\x04"


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        s = str(value).replace(",", ".").strip()
        if s == "":
            return None
        return Decimal(s)
    except (ValueError, TypeError, InvalidOperation):
        return None


def _safe_str(value: Any, max_len: int | None = None) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    if max_len is not None:
        s = s[:max_len]
    return s


class BcuUploadPreviewRow:
    def __init__(self, row_number: int, data: dict, errors: list[str] | None = None):
        self.row_number = row_number
        self.data = data
        self.errors = errors or []


class BcuUploadResult:
    def __init__(
        self,
        tipo: str,
        total_rows: int,
        valid_rows: int,
        invalid_rows: int,
        rows: list[BcuUploadPreviewRow],
        db_items: list[Any],
        base_tcpo_items: list[BaseTcpo],
    ):
        self.tipo = tipo
        self.total_rows = total_rows
        self.valid_rows = valid_rows
        self.invalid_rows = invalid_rows
        self.rows = rows
        self.db_items = db_items
        self.base_tcpo_items = base_tcpo_items


def _parse_mo(file_bytes: bytes) -> BcuUploadResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[BcuUploadPreviewRow] = []
    db_items: list[BcuMaoObraItem] = []
    base_tcpo_items: list[BaseTcpo] = []
    valid = 0
    invalid = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        errors: list[str] = []
        desc = _safe_str(row[1] if len(row) > 1 else None)
        if desc is None:
            continue

        salario = _to_decimal(row[2]) if len(row) > 2 else None
        custo_unitario = salario

        data = {
            "descricao_funcao": desc,
            "salario": salario,
            "previsao_reajuste": _to_decimal(row[3]) if len(row) > 3 else None,
            "periculosidade_insalubridade": _to_decimal(row[4]) if len(row) > 4 else None,
            "refeicao": _to_decimal(row[5]) if len(row) > 5 else None,
            "agua_potavel": _to_decimal(row[6]) if len(row) > 6 else None,
            "vale_alimentacao": _to_decimal(row[7]) if len(row) > 7 else None,
            "plano_saude": _to_decimal(row[8]) if len(row) > 8 else None,
            "seguro_vida": _to_decimal(row[9]) if len(row) > 9 else None,
            "abono_ferias": _to_decimal(row[10]) if len(row) > 10 else None,
        }

        if not desc or len(desc) > 255:
            errors.append("descricao_funcao é obrigatória e deve ter até 255 caracteres.")

        if errors:
            invalid += 1
        else:
            valid += 1
            db_items.append(
                BcuMaoObraItem(
                    id=uuid.uuid4(),
                    cabecalho_id=uuid.uuid4(),  # placeholder
                    descricao_funcao=desc,
                    **{k: v for k, v in data.items() if k != "descricao_funcao"},
                )
            )
            base_tcpo_items.append(
                BaseTcpo(
                    id=uuid.uuid4(),
                    codigo_origem=f"PLACEHOLDER-MO-{idx}",
                    descricao=desc,
                    unidade_medida="H",
                    custo_base=float(custo_unitario or 0),
                    tipo_recurso="MO",
                )
            )

        rows.append(BcuUploadPreviewRow(row_number=idx, data=data, errors=errors if errors else None))

    return BcuUploadResult(
        tipo="mo", total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        rows=rows, db_items=db_items, base_tcpo_items=base_tcpo_items,
    )


def _parse_equipamentos(file_bytes: bytes) -> BcuUploadResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[BcuUploadPreviewRow] = []
    db_items: list[BcuEquipamentoItem] = []
    base_tcpo_items: list[BaseTcpo] = []
    valid = 0
    invalid = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        errors: list[str] = []
        codigo = _safe_str(row[0] if len(row) > 0 else None, 80)
        equipamento = _safe_str(row[1] if len(row) > 1 else None)
        if equipamento is None:
            continue

        aluguel = _to_decimal(row[4]) if len(row) > 4 else None
        data = {
            "codigo": codigo,
            "equipamento": equipamento,
            "combustivel_utilizado": _safe_str(row[2] if len(row) > 2 else None, 60),
            "consumo_l_h": _to_decimal(row[3]) if len(row) > 3 else None,
            "aluguel_r_h": aluguel,
            "aluguel_mensal": _to_decimal(row[5]) if len(row) > 5 else None,
        }

        if not equipamento or len(equipamento) > 255:
            errors.append("equipamento é obrigatório e deve ter até 255 caracteres.")

        if errors:
            invalid += 1
        else:
            valid += 1
            db_items.append(
                BcuEquipamentoItem(
                    id=uuid.uuid4(),
                    cabecalho_id=uuid.uuid4(),
                    equipamento=equipamento,
                    codigo=codigo,
                    combustivel_utilizado=data["combustivel_utilizado"],
                    consumo_l_h=data["consumo_l_h"],
                    aluguel_r_h=data["aluguel_r_h"],
                    aluguel_mensal=data["aluguel_mensal"],
                )
            )
            base_tcpo_items.append(
                BaseTcpo(
                    id=uuid.uuid4(),
                    codigo_origem=f"PLACEHOLDER-EQP-{idx}",
                    descricao=equipamento,
                    unidade_medida="H",
                    custo_base=float(aluguel or 0),
                    tipo_recurso="EQUIPAMENTO",
                )
            )

        rows.append(BcuUploadPreviewRow(row_number=idx, data=data, errors=errors if errors else None))

    return BcuUploadResult(
        tipo="equipamentos", total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        rows=rows, db_items=db_items, base_tcpo_items=base_tcpo_items,
    )


def _parse_encargos(file_bytes: bytes) -> BcuUploadResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[BcuUploadPreviewRow] = []
    db_items: list[BcuEncargoItem] = []
    valid = 0
    invalid = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        errors: list[str] = []
        tipo_encargo = _safe_str(row[1] if len(row) > 1 else None, 20)
        discriminacao = _safe_str(row[4] if len(row) > 4 else None)
        if discriminacao is None:
            continue

        if tipo_encargo is None:
            tipo_encargo = "HORISTA"
        tipo_upper = tipo_encargo.upper()
        if tipo_upper not in ("HORISTA", "MENSALISTA"):
            tipo_upper = "HORISTA"

        data = {
            "tipo_encargo": tipo_upper,
            "grupo": _safe_str(row[2] if len(row) > 2 else None, 80),
            "codigo_grupo": _safe_str(row[3] if len(row) > 3 else None, 255),
            "discriminacao_encargo": discriminacao,
            "taxa_percent": _to_decimal(row[5]) if len(row) > 5 else None,
        }

        if not discriminacao or len(discriminacao) > 255:
            errors.append("discriminacao_encargo é obrigatória e deve ter até 255 caracteres.")

        if errors:
            invalid += 1
        else:
            valid += 1
            db_items.append(
                BcuEncargoItem(
                    id=uuid.uuid4(),
                    cabecalho_id=uuid.uuid4(),
                    **data,
                )
            )

        rows.append(BcuUploadPreviewRow(row_number=idx, data=data, errors=errors if errors else None))

    return BcuUploadResult(
        tipo="encargos", total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        rows=rows, db_items=db_items, base_tcpo_items=[],
    )


def _parse_epi(file_bytes: bytes) -> BcuUploadResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[BcuUploadPreviewRow] = []
    db_items: list[BcuEpiItem] = []
    base_tcpo_items: list[BaseTcpo] = []
    valid = 0
    invalid = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        errors: list[str] = []
        epi = _safe_str(row[1] if len(row) > 1 else None)
        if epi is None:
            continue

        unidade = _safe_str(row[2] if len(row) > 2 else None, 30) or "UN"
        custo_unitario = _to_decimal(row[3]) if len(row) > 3 else None
        data = {
            "epi": epi,
            "unidade": unidade,
            "custo_unitario": custo_unitario,
            "vida_util_meses": _to_decimal(row[4]) if len(row) > 4 else None,
        }

        if not epi or len(epi) > 255:
            errors.append("epi é obrigatório e deve ter até 255 caracteres.")

        if errors:
            invalid += 1
        else:
            valid += 1
            db_items.append(
                BcuEpiItem(
                    id=uuid.uuid4(),
                    cabecalho_id=uuid.uuid4(),
                    epi=epi,
                    unidade=unidade,
                    custo_unitario=custo_unitario,
                    vida_util_meses=data["vida_util_meses"],
                )
            )
            base_tcpo_items.append(
                BaseTcpo(
                    id=uuid.uuid4(),
                    codigo_origem=f"PLACEHOLDER-EPI-{idx}",
                    descricao=epi,
                    unidade_medida=unidade,
                    custo_base=float(custo_unitario or 0),
                    tipo_recurso="INSUMO",
                )
            )

        rows.append(BcuUploadPreviewRow(row_number=idx, data=data, errors=errors if errors else None))

    return BcuUploadResult(
        tipo="epi", total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        rows=rows, db_items=db_items, base_tcpo_items=base_tcpo_items,
    )


def _parse_ferramentas(file_bytes: bytes) -> BcuUploadResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[BcuUploadPreviewRow] = []
    db_items: list[BcuFerramentaItem] = []
    base_tcpo_items: list[BaseTcpo] = []
    valid = 0
    invalid = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        errors: list[str] = []
        descricao = _safe_str(row[1] if len(row) > 1 else None)
        if descricao is None:
            continue

        unidade = _safe_str(row[2] if len(row) > 2 else None, 30) or "UN"
        preco = _to_decimal(row[3]) if len(row) > 3 else None
        data = {
            "descricao": descricao,
            "unidade": unidade,
            "preco": preco,
        }

        if not descricao or len(descricao) > 255:
            errors.append("descricao é obrigatória e deve ter até 255 caracteres.")

        if errors:
            invalid += 1
        else:
            valid += 1
            db_items.append(
                BcuFerramentaItem(
                    id=uuid.uuid4(),
                    cabecalho_id=uuid.uuid4(),
                    descricao=descricao,
                    unidade=unidade,
                    preco=preco,
                )
            )
            base_tcpo_items.append(
                BaseTcpo(
                    id=uuid.uuid4(),
                    codigo_origem=f"PLACEHOLDER-FER-{idx}",
                    descricao=descricao,
                    unidade_medida=unidade,
                    custo_base=float(preco or 0),
                    tipo_recurso="FERRAMENTA",
                )
            )

        rows.append(BcuUploadPreviewRow(row_number=idx, data=data, errors=errors if errors else None))

    return BcuUploadResult(
        tipo="ferramentas", total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        rows=rows, db_items=db_items, base_tcpo_items=base_tcpo_items,
    )


def _parse_mobilizacao(file_bytes: bytes) -> BcuUploadResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[BcuUploadPreviewRow] = []
    db_items: list[BcuMobilizacaoItem] = []
    valid = 0
    invalid = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        errors: list[str] = []
        descricao = _safe_str(row[0] if len(row) > 0 else None)
        if descricao is None:
            continue

        data = {
            "descricao": descricao,
            "funcao": _safe_str(row[1] if len(row) > 1 else None, 120),
            "tipo_mao_obra": _safe_str(row[2] if len(row) > 2 else None, 20),
        }

        if not descricao or len(descricao) > 255:
            errors.append("descricao é obrigatória e deve ter até 255 caracteres.")

        if errors:
            invalid += 1
        else:
            valid += 1
            db_items.append(
                BcuMobilizacaoItem(
                    id=uuid.uuid4(),
                    cabecalho_id=uuid.uuid4(),
                    descricao=descricao,
                    funcao=data["funcao"],
                    tipo_mao_obra=data["tipo_mao_obra"],
                )
            )

        rows.append(BcuUploadPreviewRow(row_number=idx, data=data, errors=errors if errors else None))

    return BcuUploadResult(
        tipo="mobilizacao", total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        rows=rows, db_items=db_items, base_tcpo_items=[],
    )


def _parse_exames(file_bytes: bytes) -> BcuUploadResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[BcuUploadPreviewRow] = []
    base_tcpo_items: list[BaseTcpo] = []
    valid = 0
    invalid = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        errors: list[str] = []
        codigo = _safe_str(row[0] if len(row) > 0 else None, 80)
        exame = _safe_str(row[1] if len(row) > 1 else None)
        custo = _to_decimal(row[2]) if len(row) > 2 else None

        if exame is None:
            continue

        data = {"codigo": codigo, "exame": exame, "custo_unitario": custo}
        if not exame or len(exame) > 255:
            errors.append("exame é obrigatório e deve ter até 255 caracteres.")

        if errors:
            invalid += 1
        else:
            valid += 1
            base_tcpo_items.append(
                BaseTcpo(
                    id=uuid.uuid4(),
                    codigo_origem=codigo or f"PLACEHOLDER-EXM-{idx}",
                    descricao=exame,
                    unidade_medida="UN",
                    custo_base=float(custo or 0),
                    tipo_recurso="EXAMES",
                )
            )

        rows.append(BcuUploadPreviewRow(row_number=idx, data=data, errors=errors if errors else None))

    return BcuUploadResult(
        tipo="exames", total_rows=len(rows), valid_rows=valid, invalid_rows=invalid,
        rows=rows, db_items=[], base_tcpo_items=base_tcpo_items,
    )


_PARSERS = {
    "mo": _parse_mo,
    "equipamentos": _parse_equipamentos,
    "encargos": _parse_encargos,
    "exames": _parse_exames,
    "epi": _parse_epi,
    "ferramentas": _parse_ferramentas,
    "mobilizacao": _parse_mobilizacao,
}


class BcuUploadService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def preview(self, tipo: str, file_bytes: bytes, nome_arquivo: str) -> BcuUploadResult:
        tipo = tipo.lower()
        if tipo not in _VALID_TIPOS:
            raise ValidationError(
                f"Tipo '{tipo}' não é válido. Use: {', '.join(sorted(_VALID_TIPOS))}"
            )
        if not nome_arquivo.lower().endswith(".xlsx"):
            raise ValidationError("Somente arquivos .xlsx são suportados.")
        if not file_bytes:
            raise ValidationError("Arquivo vazio.")
        if len(file_bytes) > _MAX_FILE_SIZE:
            raise ValidationError(f"Arquivo excede o limite de {_MAX_FILE_SIZE // (1024*1024)}MB.")
        if not file_bytes[:4].startswith(_XLSX_MAGIC):
            raise ValidationError("Arquivo não é um XLSX válido.")

        try:
            result = _PARSERS[tipo](file_bytes)
        except Exception as exc:
            logger.error("bcu.upload.preview_failed", tipo=tipo, error=str(exc), exc_info=True)
            raise ValidationError(
                f"Falha ao processar arquivo para '{tipo}'. Verifique se as colunas estão corretas. "
                f"Detalhe: {exc.__class__.__name__}: {exc}"
            ) from exc

        return result

    async def importar(
        self,
        tipo: str,
        file_bytes: bytes,
        nome_arquivo: str,
        cabecalho_id: uuid.UUID,
        criador_id: uuid.UUID,
    ) -> BcuUploadResult:
        tipo = tipo.lower()
        cab = await self.db.get(BcuCabecalho, cabecalho_id)
        if not cab:
            raise NotFoundError("BCU cabecalho", str(cabecalho_id))

        result = await self.preview(tipo, file_bytes, nome_arquivo)
        if result.invalid_rows > 0:
            raise UnprocessableEntityError(
                f"Arquivo contém {result.invalid_rows} linha(s) inválida(s). Corrija antes de importar.",
                details={"invalid_rows": result.invalid_rows, "total_rows": result.total_rows},
            )

        seq = await self._get_next_sequence(tipo, cabecalho_id)

        if tipo == "exames":
            for item in result.base_tcpo_items:
                if not item.codigo_origem or item.codigo_origem.startswith("PLACEHOLDER-"):
                    seq += 1
                    item.codigo_origem = f"BCU-EXM-{seq:03d}"
        else:
            for idx, item in enumerate(result.db_items):
                item.cabecalho_id = cabecalho_id
                if hasattr(item, "codigo_origem"):
                    seq += 1
                    prefix = {"mo": "MO", "equipamentos": "EQP", "epi": "EPI", "ferramentas": "FER"}.get(tipo, tipo.upper()[:3])
                    item.codigo_origem = f"BCU-{prefix}-{seq:03d}"
                    # Update corresponding base_tcpo item
                    if idx < len(result.base_tcpo_items):
                        result.base_tcpo_items[idx].codigo_origem = item.codigo_origem

        self.db.add_all(result.db_items)
        await self.db.flush()

        # sync base_tcpo
        if result.base_tcpo_items:
            await self._sync_base_tcpo_batch(result.base_tcpo_items)

        logger.info(
            "bcu.upload_individual.complete",
            tipo=tipo,
            cabecalho_id=str(cabecalho_id),
            rows=result.valid_rows,
        )
        return result

    async def _get_next_sequence(self, tipo: str, cabecalho_id: uuid.UUID) -> int:
        from sqlalchemy import func, select
        if tipo == "exames":
            result = await self.db.execute(
                select(func.count(BaseTcpo.id)).where(BaseTcpo.tipo_recurso == "EXAMES")
            )
            return result.scalar() or 0
        model_map = {
            "mo": BcuMaoObraItem,
            "equipamentos": BcuEquipamentoItem,
            "encargos": BcuEncargoItem,
            "epi": BcuEpiItem,
            "ferramentas": BcuFerramentaItem,
            "mobilizacao": BcuMobilizacaoItem,
        }
        model = model_map[tipo]
        result = await self.db.execute(
            select(func.count(model.id)).where(model.cabecalho_id == cabecalho_id)
        )
        return result.scalar() or 0

    async def _sync_base_tcpo_batch(self, items: list[BaseTcpo]) -> None:
        from sqlalchemy.dialects.postgresql import insert as pg_insert

        if not items:
            return
        rows = [
            {
                "id": bt.id,
                "codigo_origem": bt.codigo_origem,
                "descricao": bt.descricao,
                "unidade_medida": bt.unidade_medida,
                "custo_base": bt.custo_base,
                "tipo_recurso": bt.tipo_recurso,
            }
            for bt in items
        ]
        stmt = pg_insert(BaseTcpo).values(rows)
        stmt = stmt.on_conflict_do_update(
            index_elements=["codigo_origem"],
            set_={
                "descricao": stmt.excluded.descricao,
                "unidade_medida": stmt.excluded.unidade_medida,
                "custo_base": stmt.excluded.custo_base,
                "tipo_recurso": stmt.excluded.tipo_recurso,
            },
        )
        await self.db.execute(stmt)
