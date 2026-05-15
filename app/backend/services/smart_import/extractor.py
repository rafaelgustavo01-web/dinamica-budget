from __future__ import annotations

import csv
import io
from dataclasses import dataclass

import openpyxl

from backend.core.exceptions import ValidationError

_MAX_FILE_SIZE = 10 * 1024 * 1024
_XLSX_MAGIC = b"PK\x03\x04"
_SUPPORTED = {"xlsx", "csv"}
_MAX_ROWS = 5000
_MAX_COLUMNS = 80


@dataclass
class SheetData:
    sheet_name: str
    rows: list[list]  # each row is a list of Python scalars (str/int/float/None)


class FileExtractor:
    @staticmethod
    def from_bytes(filename: str, content: bytes, sheet_name: str | None = None) -> SheetData:
        if len(content) > _MAX_FILE_SIZE:
            raise ValidationError(f"Arquivo excede o limite de {_MAX_FILE_SIZE // (1024 * 1024)}MB.")

        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in _SUPPORTED:
            raise ValidationError(f"Extensão .{ext} não suportada. Use xlsx ou csv.")

        if ext == "xlsx":
            return FileExtractor._parse_xlsx(content, sheet_name)
        return FileExtractor._parse_csv(content)

    @staticmethod
    def _parse_xlsx(content: bytes, sheet_name: str | None) -> SheetData:
        if not content[:4].startswith(_XLSX_MAGIC):
            raise ValidationError("Arquivo não é um XLSX válido.")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name not in wb.sheetnames:
                raise ValidationError(f"Aba '{sheet_name}' não encontrada no XLSX.")
            ws = wb[sheet_name] if sheet_name else wb.active
            name = ws.title or "Sheet1"
            rows = []
            for row_number, raw_row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_number > _MAX_ROWS:
                    raise ValidationError(f"Planilha excede o limite de {_MAX_ROWS} linhas.")
                row = list(raw_row)
                # strip trailing None columns
                while row and row[-1] is None:
                    row.pop()
                if len(row) > _MAX_COLUMNS:
                    raise ValidationError(f"Planilha excede o limite de {_MAX_COLUMNS} colunas.")
                rows.append(row)
            return SheetData(sheet_name=name, rows=rows)
        finally:
            wb.close()

    @staticmethod
    def _parse_csv(content: bytes) -> SheetData:
        text = FileExtractor._decode_csv(content)
        reader = csv.reader(io.StringIO(text))
        rows = []
        for row_number, row in enumerate(reader, start=1):
            if row_number > _MAX_ROWS:
                raise ValidationError(f"Planilha excede o limite de {_MAX_ROWS} linhas.")
            if any(c.strip() for c in row):
                if len(row) > _MAX_COLUMNS:
                    raise ValidationError(f"Planilha excede o limite de {_MAX_COLUMNS} colunas.")
                rows.append(list(row))
        return SheetData(sheet_name="Sheet1", rows=rows)

    @staticmethod
    def _decode_csv(content: bytes) -> str:
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return content.decode(enc)
            except UnicodeDecodeError:
                continue
        raise ValidationError("Não foi possível decodificar o CSV.")
