"""ETL service: parses BCU XLSX and persists all sheets to the bcu schema + syncs base_tcpo."""

from __future__ import annotations

import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy import select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.logging import get_logger
from backend.models.base_tcpo import BaseTcpo
from backend.models.bcu import (
    BcuCabecalho,
    BcuEncargoItem,
    BcuEquipamentoItem,
    BcuEquipamentoPremissa,
    BcuEpiDistribuicaoFuncao,
    BcuEpiItem,
    BcuFerramentaItem,
    BcuMaoObraItem,
    BcuMobilizacaoItem,
    BcuMobilizacaoQuantidadeFuncao,
)

logger = get_logger(__name__)


def _to_decimal(value: Any):
    """Safe cast to float-compatible Decimal; returns None on blank/error."""
    if value is None:
        return None
    try:
        return float(str(value).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def __safe_get(row: tuple | None, idx: int):
    """Retorna row[idx] se existir, caso contrário None (acesso seguro a colunas)."""
    try:
        if not row:
            return None
        return row[idx] if idx < len(row) else None
    except Exception:
        return None


def _find_col(col_map: dict[str, int], *keywords: str) -> int | None:
    """Finds first column index matching any of the keywords (case-insensitive)."""
    for kw in keywords:
        kw_up = kw.upper()
        for hk, idx in col_map.items():
            if kw_up in hk:
                return idx
    return None


def _parse_mao_obra(ws, cabecalho_id: uuid.UUID, seq_counter: dict[str, int]) -> tuple[list[BcuMaoObraItem], list[BaseTcpo]]:
    header_row_idx = 3
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if any(cell and "DESCRI" in str(cell).upper() for cell in row):
            header_row_idx = row_num
            break

    header_vals = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    col_map = {str(c).strip().upper(): i for i, c in enumerate(header_vals) if c}

    c_desc = _find_col(col_map, "DESCRI") if col_map else None
    if c_desc is None:
        c_desc = 0
    c_qtd = _find_col(col_map, "QUANT") or 1
    c_sal = _find_col(col_map, "SALARIO", "SALÁRIO") or 2
    c_reaj = _find_col(col_map, "REAJUSTE") or 3
    c_enc = _find_col(col_map, "ENCARGO") or 4
    c_peric = _find_col(col_map, "PERICULOSIDADE", "INSALUBRIDADE") or 5
    c_refei = _find_col(col_map, "REFEIÇÃO", "REFEICAO") or 6
    c_agua = _find_col(col_map, "ÁGUA", "AGUA") or 7
    c_vale = _find_col(col_map, "VALE ALIMENTAÇÃO", "VALE ALIMENTACAO") or 8
    c_saude = _find_col(col_map, "SAÚDE", "SAUDE") or 9
    c_ferr = _find_col(col_map, "FERRAMENTA") or 10
    c_seg = _find_col(col_map, "SEGURO") or 11
    c_ferias = _find_col(col_map, "FÉRIAS", "FERIAS") or 12
    c_unif = _find_col(col_map, "UNIFORME") or 13
    c_epi = _find_col(col_map, "EPI") or 14
    c_cunit = _find_col(col_map, "CUSTO UNITARIO", "UNITÁRIO") or 15
    c_cmensal = _find_col(col_map, "CUSTO MENSAL") or 16
    c_mob = _find_col(col_map, "MOBILIZAÇÃO", "MOBILIZACAO") or 18

    def _v(row: tuple, c: int | None) -> float | None:
        return _to_decimal(row[c]) if c is not None and c < len(row) else None

    items: list[BcuMaoObraItem] = []
    base_tcpo_items: list[BaseTcpo] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        desc = row[c_desc] if c_desc < len(row) else None
        if not desc:
            continue
        seq_counter["MO"] = seq_counter.get("MO", 0) + 1
        codigo_origem = f"BCU-MO-{seq_counter['MO']:03d}"
        items.append(
            BcuMaoObraItem(
                id=uuid.uuid4(),
                cabecalho_id=cabecalho_id,
                descricao_funcao=str(desc).strip(),
                codigo_origem=codigo_origem,
                quantidade=_v(row, c_qtd),
                salario=_v(row, c_sal),
                previsao_reajuste=_v(row, c_reaj),
                encargos_percent=_v(row, c_enc),
                periculosidade_insalubridade=_v(row, c_peric),
                refeicao=_v(row, c_refei),
                agua_potavel=_v(row, c_agua),
                vale_alimentacao=_v(row, c_vale),
                plano_saude=_v(row, c_saude),
                ferramentas_val=_v(row, c_ferr),
                seguro_vida=_v(row, c_seg),
                abono_ferias=_v(row, c_ferias),
                uniforme_val=_v(row, c_unif),
                epi_val=_v(row, c_epi),
                custo_unitario_h=_v(row, c_cunit),
                custo_mensal=_v(row, c_cmensal),
                mobilizacao=_v(row, c_mob),
            )
        )
        base_tcpo_items.append(
            BaseTcpo(
                id=uuid.uuid4(),
                codigo_origem=codigo_origem,
                descricao=str(desc).strip(),
                unidade_medida="H",
                custo_base=_v(row, c_cunit) or 0.0,
                tipo_recurso="MO",
            )
        )
    return items, base_tcpo_items


def _parse_equipamentos(ws, cabecalho_id: uuid.UUID, seq_counter: dict[str, int]):
    premissa = None
    items: list[BcuEquipamentoItem] = []
    base_tcpo_items: list[BaseTcpo] = []

    all_rows = list(ws.iter_rows(values_only=True))

    horas_mes = None
    gasolina = None
    diesel = None
    for r in all_rows[:5]:
        if r:
            label = str(__safe_get(r, 8)).lower() if __safe_get(r, 8) else ""
            if "hora" in label:
                horas_mes = _to_decimal(__safe_get(r, 9))
            if "gasolina" in label:
                gasolina = _to_decimal(__safe_get(r, 9))
            if "diesel" in label:
                diesel = _to_decimal(__safe_get(r, 9))

    premissa = BcuEquipamentoPremissa(
        id=uuid.uuid4(),
        cabecalho_id=cabecalho_id,
        horas_mes=horas_mes,
        preco_gasolina_l=gasolina,
        preco_diesel_l=diesel,
    )

    for row in ws.iter_rows(min_row=7, values_only=True):
        equip = __safe_get(row, 1)
        if not equip:
            continue
        seq_counter["EQP"] = seq_counter.get("EQP", 0) + 1
        codigo_origem = f"BCU-EQP-{seq_counter['EQP']:03d}"
        try:
            items.append(
                BcuEquipamentoItem(
                    id=uuid.uuid4(),
                    cabecalho_id=cabecalho_id,
                    codigo=str(__safe_get(row, 0)).strip() if __safe_get(row, 0) else None,
                    codigo_origem=codigo_origem,
                    equipamento=str(equip).strip(),
                    combustivel_utilizado=str(__safe_get(row, 2)).strip() if __safe_get(row, 2) else None,
                    consumo_l_h=_to_decimal(__safe_get(row, 3)),
                    aluguel_r_h=_to_decimal(__safe_get(row, 4)),
                    combustivel_r_h=_to_decimal(__safe_get(row, 5)),
                    mao_obra_r_h=_to_decimal(__safe_get(row, 6)),
                    hora_produtiva=_to_decimal(__safe_get(row, 7)),
                    hora_improdutiva=_to_decimal(__safe_get(row, 8)),
                    mes=_to_decimal(__safe_get(row, 9)),
                    aluguel_mensal=_to_decimal(__safe_get(row, 11)),
                )
            )
        except Exception as e:
            logger.exception("bcu.equipamento_row_failed", sheet=str(ws.title), row_preview=str(row))
            continue
        try:
            base_tcpo_items.append(
                BaseTcpo(
                    id=uuid.uuid4(),
                    codigo_origem=codigo_origem,
                    descricao=str(equip).strip(),
                    unidade_medida="H",
                    custo_base=_to_decimal(__safe_get(row, 4)) or 0.0,
                    tipo_recurso="EQUIPAMENTO",
                )
            )
        except Exception:
            logger.exception("bcu.base_tcpo_equipamento_failed", sheet=str(ws.title), row_preview=str(row))
            continue
    return premissa, items, base_tcpo_items


def _parse_encargos(ws, cabecalho_id: uuid.UUID, tipo: str) -> list[BcuEncargoItem]:
    items: list[BcuEncargoItem] = []
    current_grupo = None

    for row in ws.iter_rows(min_row=4, values_only=True):
        if row and len(row) > 0 and row[0]:
            try:
                val0 = str(row[0]).strip()
            except Exception:
                val0 = None
            if val0 and len(val0) <= 3 and val0.isalpha():
                current_grupo = val0

        if tipo == "HORISTA":
            cod = str(row[1]).strip() if row and len(row) > 1 and row[1] else None
            disc = str(row[2]).strip() if row and len(row) > 2 and row[2] else None
            taxa = _to_decimal(row[5]) if row and len(row) > 5 else None
        else:
            cod = str(row[1]).strip() if row and len(row) > 1 and row[1] else None
            disc = str(row[2]).strip() if row and len(row) > 2 and row[2] else None
            taxa = _to_decimal(row[3]) if row and len(row) > 3 else None

        if not disc or disc.lower() in ("discriminação do encargo", "grupos"):
            continue
        if disc.lower().startswith("total do grupo"):
            continue

        items.append(
            BcuEncargoItem(
                id=uuid.uuid4(),
                cabecalho_id=cabecalho_id,
                tipo_encargo=tipo,
                grupo=current_grupo,
                codigo_grupo=cod,
                discriminacao_encargo=disc,
                taxa_percent=taxa,
            )
        )
    return items


_EPI_FUNCOES = [
    "OFICIAL", "AJUDANTE", "ELETRICISTA", "OPERADOR",
    "MÃO DE OBRA INDIRETA", "ALPINISTA", "MONTADOR ANDAIME",
]


def _parse_epi(ws, cabecalho_id: uuid.UUID, seq_counter: dict[str, int]):
    epi_items: list[BcuEpiItem] = []
    dist_items: list[BcuEpiDistribuicaoFuncao] = []
    base_tcpo_items: list[BaseTcpo] = []
    header_rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    header_row = header_rows[0] if header_rows else []
    funcao_cols: list[tuple[int, str]] = []
    for idx, cell in enumerate(header_row):
        if cell and idx >= 8:
            try:
                funcao_cols.append((idx, str(cell).strip()))
            except Exception:
                continue

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) <= 1:
            continue
        epi_name = __safe_get(row, 1)
        if not epi_name:
            continue

        seq_counter["EPI"] = seq_counter.get("EPI", 0) + 1
        codigo_origem = f"BCU-EPI-{seq_counter['EPI']:03d}"
        epi_id = uuid.uuid4()
        try:
            epi_items.append(
                BcuEpiItem(
                    id=epi_id,
                    cabecalho_id=cabecalho_id,
                    codigo_origem=codigo_origem,
                    epi=str(epi_name).strip(),
                    unidade=str(__safe_get(row, 2)).strip() if __safe_get(row, 2) else None,
                    custo_unitario=_to_decimal(__safe_get(row, 3)),
                    quantidade=_to_decimal(__safe_get(row, 4)),
                    vida_util_meses=_to_decimal(__safe_get(row, 5)),
                    custo_epi_mes=_to_decimal(__safe_get(row, 6)),
                )
            )
        except Exception as e:
            logger.exception("bcu.epi_row_failed", sheet=str(ws.title), row_preview=str(row), error=str(e))
            continue
        base_tcpo_items.append(
            BaseTcpo(
                id=uuid.uuid4(),
                codigo_origem=codigo_origem,
                descricao=str(epi_name).strip(),
                unidade_medida=str(__safe_get(row, 2)).strip() if __safe_get(row, 2) else "UN",
                custo_base=(_to_decimal(__safe_get(row, 3)) or 0.0),
                tipo_recurso="INSUMO",
            )
        )

        for col_idx, funcao in funcao_cols:
            val = __safe_get(row, col_idx)
            if val is not None and str(val).strip() != "":
                dist_items.append(
                    BcuEpiDistribuicaoFuncao(
                        id=uuid.uuid4(),
                        epi_item_id=epi_id,
                        funcao=funcao,
                        aplica_flag=str(val).strip(),
                    )
                )

    return epi_items, dist_items, base_tcpo_items


def _parse_ferramentas(ws, cabecalho_id: uuid.UUID, seq_counter: dict[str, int]) -> tuple[list[BcuFerramentaItem], list[BaseTcpo]]:
    items: list[BcuFerramentaItem] = []
    base_tcpo_items: list[BaseTcpo] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        desc = row[2] if len(row) > 2 else None
        if not desc:
            continue
        seq_counter["FER"] = seq_counter.get("FER", 0) + 1
        codigo_origem = f"BCU-FER-{seq_counter['FER']:03d}"
        items.append(
            BcuFerramentaItem(
                id=uuid.uuid4(),
                cabecalho_id=cabecalho_id,
                codigo_origem=codigo_origem,
                item=str(row[1]).strip() if row[1] else None,
                descricao=str(desc).strip(),
                unidade=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                quantidade=_to_decimal(row[4]) if len(row) > 4 else None,
                preco=_to_decimal(row[5]) if len(row) > 5 else None,
                preco_total=_to_decimal(row[6]) if len(row) > 6 else None,
            )
        )
        base_tcpo_items.append(
            BaseTcpo(
                id=uuid.uuid4(),
                codigo_origem=codigo_origem,
                descricao=str(desc).strip(),
                unidade_medida=str(row[3]).strip() if len(row) > 3 and row[3] else "UN",
                custo_base=_to_decimal(row[5]) or 0.0,
                tipo_recurso="FERRAMENTA",
            )
        )
    return items, base_tcpo_items


def _parse_mobilizacao(ws, cabecalho_id: uuid.UUID):
    mob_items: list[BcuMobilizacaoItem] = []
    quant_items: list[BcuMobilizacaoQuantidadeFuncao] = []

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    header_row_idx = 1
    for i, row in enumerate(all_rows[:6]):
        if row and len(row) > 2 and any(cell for cell in row[2:] if cell):
            header_row_idx = i
            break

    header_row2 = all_rows[header_row_idx] if len(all_rows) > header_row_idx else []
    funcao_cols: list[tuple[int, str]] = []
    for idx, cell in enumerate(header_row2):
        if cell and idx >= 2:
            funcao_cols.append((idx, str(cell).strip()))

    data_start_idx = None
    for i in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[i]
        if row and row[0]:
            data_start_idx = i
            break

    if data_start_idx is None:
        data_start_idx = len(all_rows)

    for row in all_rows[data_start_idx:]:
        desc = row[0]
        if not desc:
            continue
        item_id = uuid.uuid4()
        valor_unitario = row[1] if len(row) > 1 else None
        mob_items.append(
            BcuMobilizacaoItem(
                id=item_id,
                cabecalho_id=cabecalho_id,
                descricao=str(desc).strip(),
                funcao=str(valor_unitario).strip() if valor_unitario else None,
                tipo_mao_obra=None,
            )
        )
        for col_idx, funcao in funcao_cols:
            val = row[col_idx] if len(row) > col_idx else None
            quant_items.append(
                BcuMobilizacaoQuantidadeFuncao(
                    id=uuid.uuid4(),
                    mobilizacao_item_id=item_id,
                    coluna_funcao=funcao,
                    quantidade=_to_decimal(val),
                )
            )

    return mob_items, quant_items


class BcuService:
    """
    Importa o arquivo BCU.xlsx (planilha mestra de custos) com 7 abas.
    Ao importar:
      1. Cria bcu.cabecalho (is_ativo=False inicialmente)
      2. Popula 9 tabelas filhas
      3. Sincroniza referencia.base_tcpo:
         - Para cada item de MO/EQP/EPI/FER, cria/atualiza BaseTcpo
         - codigo_origem = BCU-{tipo}-{N}
      4. Encargos e Mobilizacao NAO sao sincronizados
    """

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def importar_bcu(
        self,
        file_bytes: bytes,
        nome_arquivo: str,
        criador_id: uuid.UUID,
    ) -> BcuCabecalho:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

        # Remove previous data with same filename
        result = await self.db.execute(select(BcuCabecalho).where(BcuCabecalho.nome_arquivo == nome_arquivo))
        existing = result.scalars().all()
        for cab in existing:
            await self.db.delete(cab)
        await self.db.flush()

        cab = BcuCabecalho(
            id=uuid.uuid4(),
            nome_arquivo=nome_arquivo,
            versao_layout="v1",
            is_ativo=False,
            criado_por_id=criador_id,
        )
        self.db.add(cab)
        await self.db.flush()

        seq_counter: dict[str, int] = {}
        all_base_tcpo: list[BaseTcpo] = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            if "MÃO DE OBRA" in sheet_name.upper() and "INDIRETA" not in sheet_name.upper():
                ws = wb[sheet_name]
                try:
                    items, bt_items = _parse_mao_obra(ws, cab.id, seq_counter)
                except Exception as e:
                    logger.exception("bcu.parse_mao_obra_failed", sheet=sheet_name)
                    raise ValueError(f"Erro ao parsear MÃO DE OBRA ({sheet_name}): {e}") from e
                for item in items:
                    self.db.add(item)
                    total_rows += 1
                all_base_tcpo.extend(bt_items)
                break

        for sheet_name in wb.sheetnames:
            if "EQUIPAMENTO" in sheet_name.upper():
                ws = wb[sheet_name]
                try:
                    premissa, eq_items, bt_items = _parse_equipamentos(ws, cab.id, seq_counter)
                except Exception as e:
                    logger.exception("bcu.parse_equipamentos_failed", sheet=sheet_name)
                    raise ValueError(f"Erro ao parsear EQUIPAMENTO ({sheet_name}): {e}") from e
                self.db.add(premissa)
                for item in eq_items:
                    self.db.add(item)
                    total_rows += 1
                all_base_tcpo.extend(bt_items)
                break

        horista_done = False
        for sheet_name in wb.sheetnames:
            if "HORISTA" in sheet_name.upper():
                ws = wb[sheet_name]
                try:
                    enc_items = _parse_encargos(ws, cab.id, "HORISTA")
                except Exception as e:
                    logger.exception("bcu.parse_encargos_failed", sheet=sheet_name)
                    raise ValueError(f"Erro ao parsear ENCARGOS HORISTA ({sheet_name}): {e}") from e
                for item in enc_items:
                    self.db.add(item)
                    total_rows += 1
                horista_done = True
                break

        mensalista_done = False
        for sheet_name in wb.sheetnames:
            if "MENSALISTA" in sheet_name.upper():
                ws = wb[sheet_name]
                try:
                    enc_items = _parse_encargos(ws, cab.id, "MENSALISTA")
                except Exception as e:
                    logger.exception("bcu.parse_encargos_failed", sheet=sheet_name)
                    raise ValueError(f"Erro ao parsear ENCARGOS MENSALISTA ({sheet_name}): {e}") from e
                for item in enc_items:
                    self.db.add(item)
                    total_rows += 1
                mensalista_done = True
                break

        if not horista_done and not mensalista_done:
            for sheet_name in wb.sheetnames:
                if "ENCARGO" in sheet_name.upper():
                    ws = wb[sheet_name]
                    try:
                        enc_items = _parse_encargos(ws, cab.id, "HORISTA")
                    except Exception as e:
                        logger.exception("bcu.parse_encargos_failed", sheet=sheet_name)
                        raise ValueError(f"Erro ao parsear ENCARGOS ({sheet_name}): {e}") from e
                    for item in enc_items:
                        self.db.add(item)
                        total_rows += 1
                    break

        for sheet_name in wb.sheetnames:
            if "EPI" in sheet_name.upper():
                ws = wb[sheet_name]
                try:
                    epi_items, dist_items, bt_items = _parse_epi(ws, cab.id, seq_counter)
                except Exception as e:
                    logger.exception("bcu.parse_epi_failed", sheet=sheet_name)
                    raise ValueError(f"Erro ao parsear EPI ({sheet_name}): {e}") from e
                for item in epi_items:
                    self.db.add(item)
                    total_rows += 1
                for dist in dist_items:
                    self.db.add(dist)
                all_base_tcpo.extend(bt_items)
                break

        for sheet_name in wb.sheetnames:
            if "FERRAMENTA" in sheet_name.upper():
                ws = wb[sheet_name]
                try:
                    items, bt_items = _parse_ferramentas(ws, cab.id, seq_counter)
                except Exception as e:
                    logger.exception("bcu.parse_ferramentas_failed", sheet=sheet_name)
                    raise ValueError(f"Erro ao parsear FERRAMENTA ({sheet_name}): {e}") from e
                for item in items:
                    self.db.add(item)
                    total_rows += 1
                all_base_tcpo.extend(bt_items)
                break

        for sheet_name in wb.sheetnames:
            if "MOBILIZA" in sheet_name.upper():
                ws = wb[sheet_name]
                try:
                    mob_items, quant_items = _parse_mobilizacao(ws, cab.id)
                except Exception as e:
                    logger.exception("bcu.parse_mobilizacao_failed", sheet=sheet_name)
                    raise ValueError(f"Erro ao parsear MOBILIZACAO ({sheet_name}): {e}") from e
                for item in mob_items:
                    self.db.add(item)
                    total_rows += 1
                for q in quant_items:
                    self.db.add(q)
                break

        # Sync base_tcpo via upsert
        for bt in all_base_tcpo:
            stmt = pg_insert(BaseTcpo).values(
                id=bt.id,
                codigo_origem=bt.codigo_origem,
                descricao=bt.descricao,
                unidade_medida=bt.unidade_medida,
                custo_base=bt.custo_base,
                tipo_recurso=bt.tipo_recurso,
            )
            stmt = stmt.on_conflict_do_update(
                index_elements=["codigo_origem"],
                set_={
                    "descricao": stmt.excluded.descricao,
                    "unidade_medida": stmt.excluded.unidade_medida,
                    "custo_base": stmt.excluded.custo_base,
                    "tipo_recurso": stmt.excluded.tipo_recurso,
                },
            )
            await self.db.execute(stmt)

        logger.info(
            "bcu.import_complete",
            arquivo=nome_arquivo,
            cabecalho_id=str(cab.id),
            rows=total_rows,
            base_tcpo_synced=len(all_base_tcpo),
        )
        await self.db.commit()
        await self.db.refresh(cab)
        return cab

    async def ativar_cabecalho(self, cabecalho_id: uuid.UUID) -> BcuCabecalho:
        cab = await self.db.get(BcuCabecalho, cabecalho_id)
        if not cab:
            raise ValueError("Cabecalho nao encontrado")

        # Desativa todos
        await self.db.execute(
            update(BcuCabecalho).values(is_ativo=False)
        )
        # Ativa o solicitado
        cab.is_ativo = True
        self.db.add(cab)
        await self.db.commit()
        await self.db.refresh(cab)
        return cab

    async def get_cabecalho_ativo(self) -> BcuCabecalho | None:
        result = await self.db.execute(
            select(BcuCabecalho).where(BcuCabecalho.is_ativo == True).limit(1)
        )
        return result.scalar_one_or_none()
