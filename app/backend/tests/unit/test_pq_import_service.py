from io import BytesIO
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

import openpyxl
import pytest

from backend.core.exceptions import ValidationError
from backend.models.enums import StatusImportacao, StatusProposta
from backend.models.pq_layout import PqLayoutCliente
from backend.services.pq_import_service import PqImportService, _is_likely_section_title


def _build_upload(filename: str, payload: bytes):
    arquivo = MagicMock()
    arquivo.filename = filename
    arquivo.read = AsyncMock(return_value=payload)
    return arquivo


@pytest.mark.asyncio
async def test_importar_planilha_csv_cria_itens_e_normaliza_tokens():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta
    importacao_repo = AsyncMock()
    importacao_repo.create.side_effect = lambda imp: imp
    importacao_repo.update.side_effect = lambda imp: imp
    item_repo = AsyncMock()
    item_repo.create_batch.side_effect = lambda items: items

    svc = PqImportService(proposta_repo, importacao_repo, item_repo)
    arquivo = _build_upload(
        "pq.csv",
        b"codigo,descricao,unidade,quantidade\n001,Escavacao manual,m2,10.5\n",
    )

    resultado = await svc.importar_planilha(proposta.id, arquivo)

    assert resultado.status == StatusImportacao.CONCLUIDO
    assert resultado.linhas_total == 1
    assert resultado.linhas_importadas == 1
    assert proposta.status == StatusProposta.EM_ANALISE
    created_items = item_repo.create_batch.await_args.args[0]
    assert created_items[0].descricao_tokens == "escavacao manual"


@pytest.mark.asyncio
async def test_importar_planilha_xlsx_reconhece_aliases_de_coluna():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta
    importacao_repo = AsyncMock()
    importacao_repo.create.side_effect = lambda imp: imp
    importacao_repo.update.side_effect = lambda imp: imp
    item_repo = AsyncMock()
    item_repo.create_batch.side_effect = lambda items: items

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Código", "Descrição", "Unidade", "Quantidade"])
    ws.append(["A-01", "Concreto estrutural", "m3", 12])
    payload = BytesIO()
    wb.save(payload)

    svc = PqImportService(proposta_repo, importacao_repo, item_repo)
    arquivo = _build_upload("pq.xlsx", payload.getvalue())

    resultado = await svc.importar_planilha(proposta.id, arquivo)

    assert resultado.status == StatusImportacao.CONCLUIDO
    created_items = item_repo.create_batch.await_args.args[0]
    assert created_items[0].codigo_original == "A-01"
    assert str(created_items[0].quantidade_original) == "12"


@pytest.mark.asyncio
async def test_importar_planilha_rejeita_extensao_nao_suportada():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta
    svc = PqImportService(proposta_repo, AsyncMock(), AsyncMock())
    arquivo = _build_upload("pq.txt", b"teste")

    with pytest.raises(ValidationError, match="Somente arquivos"):
        await svc.importar_planilha(proposta.id, arquivo)


@pytest.mark.asyncio
async def test_preview_planilha_retorna_itens_sem_gravar():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta
    importacao_repo = AsyncMock()
    item_repo = AsyncMock()

    svc = PqImportService(proposta_repo, importacao_repo, item_repo)
    arquivo = _build_upload(
        "pq.csv",
        b"codigo,descricao,unidade,quantidade\n001,Escavacao manual,m2,10.5\n002,,m2,5\n",
    )

    preview = await svc.preview_planilha(proposta.id, arquivo)

    assert preview["linhas_total"] == 2
    assert preview["linhas_ok"] == 1
    assert preview["linhas_com_erro"] == 1
    assert preview["score_confianca"] == 0
    importacao_repo.create.assert_not_awaited()
    item_repo.create_batch.assert_not_awaited()


@pytest.mark.asyncio
async def test_preview_planilha_com_layout_retorna_score():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta

    layout = MagicMock(spec=PqLayoutCliente)
    layout.aba_nome = None
    layout.linha_inicio = 1
    layout.mapeamentos = []
    layout.aliases_json = None
    layout_repo = AsyncMock()
    layout_repo.get_by_cliente_id.return_value = layout

    svc = PqImportService(proposta_repo, AsyncMock(), AsyncMock(), pq_layout_repo=layout_repo)
    arquivo = _build_upload(
        "pq.csv",
        b"codigo,descricao,unidade,quantidade\n001,Escavacao manual,m2,10.5\n",
    )

    preview = await svc.preview_planilha(proposta.id, arquivo)
    assert preview["score_confianca"] == 0


@pytest.mark.asyncio
async def test_importar_planilha_respeita_linha_inicio_do_layout_csv():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta
    importacao_repo = AsyncMock()
    importacao_repo.create.side_effect = lambda imp: imp
    importacao_repo.update.side_effect = lambda imp: imp
    item_repo = AsyncMock()
    item_repo.create_batch.side_effect = lambda items: items

    layout = MagicMock(spec=PqLayoutCliente)
    layout.aba_nome = None
    layout.linha_inicio = 4
    m1 = MagicMock()
    m1.campo_sistema.value = "descricao"
    m1.coluna_planilha = "Descricao"
    layout.mapeamentos = [m1]
    layout_repo = AsyncMock()
    layout_repo.get_by_cliente_id.return_value = layout

    svc = PqImportService(proposta_repo, importacao_repo, item_repo, pq_layout_repo=layout_repo)
    arquivo = _build_upload(
        "pq.csv",
        (
            "Descricao,Unidade,Quantidade\n"
            "linha2,m2,1\n"
            "linha3,m2,1\n"
            "Escavacao,m2,10\n"
        ).encode("utf-8"),
    )

    resultado = await svc.importar_planilha(proposta.id, arquivo)
    created_items = item_repo.create_batch.await_args.args[0]
    assert len(created_items) == 1
    assert created_items[0].descricao_original == "Escavacao"


def test_is_likely_section_title_detecta_titulos_corretamente():
    # Títulos claros
    assert _is_likely_section_title({"descricao": "SERVIÇOS PRELIMINARES", "quantidade": None, "unidade": None})
    assert _is_likely_section_title({"descricao": "1.", "quantidade": "", "unidade": ""})
    assert _is_likely_section_title({"descricao": "2.1.3", "quantidade": None, "unidade": None})
    assert _is_likely_section_title({"descricao": "CAPÍTULO 1", "quantidade": "", "unidade": ""})
    assert _is_likely_section_title({"descricao": "ETAPA FUNDAÇÃO", "quantidade": None, "unidade": None})
    assert _is_likely_section_title({"descricao": "A", "quantidade": "", "unidade": ""})

    # Itens reais (não devem ser detectados como título)
    assert not _is_likely_section_title({"descricao": "Escavação manual", "quantidade": "10", "unidade": "m2"})
    assert not _is_likely_section_title({"descricao": "Concreto estrutural", "quantidade": "5.5", "unidade": "m3"})
    assert not _is_likely_section_title({"descricao": "Revestimento cerâmico", "quantidade": "0", "unidade": "m2"})
    # zero quantidade mas com unidade → ainda é item (pode ser erro de digitação, não título)
    assert not _is_likely_section_title({"descricao": "Pintura", "quantidade": "0", "unidade": "m2"})


@pytest.mark.asyncio
async def test_importar_planilha_ignora_titulos_e_conta_linhas():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta
    importacao_repo = AsyncMock()
    importacao_repo.create.side_effect = lambda imp: imp
    importacao_repo.update.side_effect = lambda imp: imp
    item_repo = AsyncMock()
    item_repo.create_batch.side_effect = lambda items: items

    svc = PqImportService(proposta_repo, importacao_repo, item_repo)
    arquivo = _build_upload(
        "pq.csv",
        (
            "codigo,descricao,unidade,quantidade\n"
            ",SERVICOS PRELIMINARES,,\n"
            "001,Escavacao manual,m2,10.5\n"
            ",1. FUNDACAO,,\n"
            "002,Concreto estrutural,m3,12\n"
        ).encode("utf-8"),
    )

    resultado = await svc.importar_planilha(proposta.id, arquivo)

    assert resultado.linhas_total == 4
    assert resultado.linhas_importadas == 2
    assert resultado.linhas_ignoradas == 2
    assert resultado.linhas_com_erro == 0
    created_items = item_repo.create_batch.await_args.args[0]
    assert len(created_items) == 2
    assert created_items[0].descricao_original == "Escavacao manual"
    assert created_items[1].descricao_original == "Concreto estrutural"


@pytest.mark.asyncio
async def test_preview_planilha_marca_titulos_como_ignorado():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta

    svc = PqImportService(proposta_repo, AsyncMock(), AsyncMock())
    arquivo = _build_upload(
        "pq.csv",
        (
            "codigo,descricao,unidade,quantidade\n"
            ",SERVICOS PRELIMINARES,,\n"
            "001,Escavacao manual,m2,10.5\n"
            ",ETAPA 2,,\n"
        ).encode("utf-8"),
    )

    preview = await svc.preview_planilha(proposta.id, arquivo)

    assert preview["linhas_total"] == 3
    assert preview["linhas_ok"] == 1
    assert preview["linhas_ignoradas"] == 2
    assert preview["linhas_com_erro"] == 0
    itens = preview["itens"]
    assert itens[0]["status"] == "IGNORADO"
    assert itens[1]["status"] == "OK"
    assert itens[2]["status"] == "IGNORADO"


@pytest.mark.asyncio
async def test_importar_planilha_respeita_aba_nome_do_layout_xlsx():
    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.status = StatusProposta.RASCUNHO

    proposta_repo = AsyncMock()
    proposta_repo.get_by_id.return_value = proposta
    importacao_repo = AsyncMock()
    importacao_repo.create.side_effect = lambda imp: imp
    importacao_repo.update.side_effect = lambda imp: imp
    item_repo = AsyncMock()
    item_repo.create_batch.side_effect = lambda items: items

    layout = MagicMock(spec=PqLayoutCliente)
    layout.aba_nome = "PQ"
    layout.linha_inicio = 1
    m1 = MagicMock()
    m1.campo_sistema.value = "descricao"
    m1.coluna_planilha = "Descricao"
    layout.mapeamentos = [m1]
    layout_repo = AsyncMock()
    layout_repo.get_by_cliente_id.return_value = layout

    wb = openpyxl.Workbook()
    wb.active.title = "Outra"
    ws = wb.create_sheet("PQ")
    ws.append(["Descricao"])
    ws.append(["Escavacao"])
    payload = BytesIO()
    wb.save(payload)

    svc = PqImportService(proposta_repo, importacao_repo, item_repo, pq_layout_repo=layout_repo)
    arquivo = _build_upload("pq.xlsx", payload.getvalue())

    resultado = await svc.importar_planilha(proposta.id, arquivo)
    created_items = item_repo.create_batch.await_args.args[0]
    assert len(created_items) == 1
    assert created_items[0].descricao_original == "Escavacao"
