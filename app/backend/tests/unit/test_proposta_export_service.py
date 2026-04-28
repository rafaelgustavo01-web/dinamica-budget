from decimal import Decimal
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from backend.services.proposta_export_service import PropostaExportService


@pytest.mark.asyncio
async def test_gerar_excel_contem_quatro_abas(monkeypatch):
    db = MagicMock()
    svc = PropostaExportService(db)

    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.codigo = "PROP-2026-0001"
    proposta.titulo = "Obra Teste"
    proposta.status.value = "CPU_GERADA"
    proposta.cliente_id = uuid4()
    proposta.total_direto = Decimal("100000.00")
    proposta.total_indireto = Decimal("28500.00")
    proposta.total_geral = Decimal("128500.00")
    proposta.descricao = None
    proposta.created_at = None
    proposta.data_finalizacao = None

    cliente = MagicMock()
    cliente.nome_fantasia = "Cliente Teste"
    cliente.cnpj = "12.345.678/0001-90"

    item = MagicMock()
    item.id = uuid4()
    item.codigo = "001"
    item.descricao = "Escavacao manual"
    item.unidade_medida = "m3"
    item.quantidade = Decimal("10")
    item.custo_direto_unitario = Decimal("100.00")
    item.custo_indireto_unitario = Decimal("28.50")
    item.preco_unitario = Decimal("128.50")
    item.preco_total = Decimal("1285.00")
    item.percentual_indireto = Decimal("28.5")

    composicao = MagicMock()
    composicao.descricao_insumo = "Pedreiro"
    composicao.unidade_medida = "h"
    composicao.quantidade_consumo = Decimal("8")
    composicao.custo_unitario_insumo = Decimal("45.00")
    composicao.custo_total_insumo = Decimal("360.00")
    composicao.tipo_recurso = MagicMock()
    composicao.tipo_recurso.value = "MO"
    composicao.nivel = 0

    svc.proposta_repo.get_by_id = AsyncMock(return_value=proposta)
    svc.cliente_repo.get_by_id = AsyncMock(return_value=cliente)
    svc.item_repo.list_by_proposta = AsyncMock(return_value=[item])
    svc.composicao_repo.list_by_proposta_item = AsyncMock(return_value=[composicao])

    raw = await svc.gerar_excel(proposta.id)

    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) == {"Capa", "Quadro-Resumo", "CPU", "Composicoes"}
    assert wb["Capa"]["B2"].value == "Cliente Teste"  # M-03 fix: B2 shows client name, not codigo
    assert wb["CPU"].max_row >= 2
    assert wb["Composicoes"].max_row >= 2


@pytest.mark.asyncio
async def test_gerar_excel_proposta_inexistente_levanta_404():
    from backend.core.exceptions import NotFoundError

    db = MagicMock()
    svc = PropostaExportService(db)
    svc.proposta_repo.get_by_id = AsyncMock(return_value=None)

    with pytest.raises(NotFoundError):
        await svc.gerar_excel(uuid4())


@pytest.mark.asyncio
async def test_gerar_pdf_retorna_bytes_pdf(monkeypatch):
    db = MagicMock()
    svc = PropostaExportService(db)

    proposta = MagicMock()
    proposta.id = uuid4()
    proposta.codigo = "PROP-2026-0001"
    proposta.titulo = "Obra Teste"
    proposta.status.value = "CPU_GERADA"
    proposta.cliente_id = uuid4()
    proposta.total_direto = Decimal("100000.00")
    proposta.total_indireto = Decimal("28500.00")
    proposta.total_geral = Decimal("128500.00")

    cliente = MagicMock()
    cliente.nome_fantasia = "Cliente Teste"
    cliente.cnpj = "12.345.678/0001-90"

    svc.proposta_repo.get_by_id = AsyncMock(return_value=proposta)
    svc.cliente_repo.get_by_id = AsyncMock(return_value=cliente)

    raw = await svc.gerar_pdf(proposta.id)
    assert raw[:4] == b"%PDF"
    assert len(raw) > 500
