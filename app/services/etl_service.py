"""
ETL service — parse TCPO Excel files and load into referencia.* schema.

Flow:
  1. Admin uploads .xlsx  → parse_tcpo_pini() or parse_converter_datacenter()
     returns EtlUploadResponse with a parse_token (UUID key into in-memory cache)
  2. Admin hits /execute  → execute_load() pops the cached results, writes DB
  3. Admin hits /status   → get_status() returns row counts

Parse is synchronous (CPU-bound Excel parsing).
DB writes are async.
"""

from __future__ import annotations

import io
import time
import uuid
from dataclasses import dataclass, field

import openpyxl
from sqlalchemy import func, select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import get_logger
from app.models.base_tcpo import BaseTcpo
from app.models.composicao_base import ComposicaoBase
from app.schemas.etl import (
    EtlExecuteRequest,
    EtlExecuteResponse,
    EtlItemPreview,
    EtlMode,
    EtlParsePreview,
    EtlRelacaoPreview,
    EtlStatusResponse,
    EtlUploadResponse,
)

logger = get_logger(__name__)

_BATCH_SIZE = 500
_SAMPLE_SIZE = 5

# Maps Excel CLASS column → TipoRecurso enum string (or None for composite services)
_CLASS_TO_TIPO: dict[str, str | None] = {
    "MAT.": "INSUMO",
    "M.O.": "MO",
    "EQP.": "EQUIPAMENTO",
    "FER.": "FERRAMENTA",
    "SER.CG": "SERVICO",
}

# Converter sheet → (code_prefix, tipo_recurso, code_col, desc_col, unit_col, price_col)
# unit_col=None means fall back to "UN" / "H"
_CONVERTER_SHEETS: dict[str, tuple[str, str, int, int, int | None, int]] = {
    "EPI-UNIFORME": ("EPI", "INSUMO",       0, 1, 2, 3),
    "EQUIPAMENTOS": ("EQP", "EQUIPAMENTO",  0, 1, None, 4),   # price = aluguel R$/h
    "EXAMES":       ("EXM", "SERVICO",      0, 1, None, 2),
    "FERRAMENTAS":  ("FER", "FERRAMENTA",   0, 1, 2, 3),
    "MÃO DE OBRA":  ("MO",  "MO",           0, 1, None, 2),   # price = salário
}


# ── Internal parse result (not a Pydantic model — stays in memory) ─────────────

@dataclass
class _ParsedItem:
    codigo_origem: str
    descricao: str
    unidade_medida: str
    custo_base: float
    tipo_recurso: str | None = None


@dataclass
class _ParsedRelacao:
    pai_codigo: str
    filho_codigo: str
    quantidade_consumo: float
    unidade_medida: str


@dataclass
class _EtlParseResult:
    itens: list[_ParsedItem] = field(default_factory=list)
    relacoes: list[_ParsedRelacao] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)


# ── Service ────────────────────────────────────────────────────────────────────

class EtlService:
    """
    Stateful service that caches parsed Excel data between upload and execute calls.
    One instance per application (module-level singleton at bottom of file).
    """

    def __init__(self) -> None:
        self._cache: dict[str, _EtlParseResult] = {}

    # ── Parse: TCPO Composições ────────────────────────────────────────────────

    def parse_tcpo_pini(self, file_bytes: bytes) -> EtlUploadResponse:
        """
        Parse 'Composições analíticas' sheet from Composições TCPO - PINI.xlsx.

        Row detection:
          - Skip if col1 (CÓDIGO) is not a str, or CLASS is None → section header
          - CLASS == 'SER.CG'  → new parent service
          - CLASS in MAT./M.O./EQP./FER. → child component of current parent
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            ws = wb["Composições analíticas"]
        except KeyError:
            wb.close()
            raise ValueError("Planilha 'Composições analíticas' não encontrada no arquivo.")

        result = _EtlParseResult()
        current_parent_codigo: str | None = None
        seen_itens: set[str] = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            codigo = row[0]
            descricao = row[1]
            classe = row[2]
            unidade = row[3]
            coef = row[4]
            preco = row[5]

            # Skip section headers: code is not a str, or CLASS is None
            if not isinstance(codigo, str) or classe is None:
                continue

            classe_clean = str(classe).strip()
            tipo_recurso = _CLASS_TO_TIPO.get(classe_clean)
            unidade_clean = str(unidade).strip() if unidade else "UN"
            descricao_clean = str(descricao).strip() if descricao else ""
            custo = float(preco) if preco is not None else 0.0

            if classe_clean == "SER.CG":
                current_parent_codigo = codigo
                if codigo not in seen_itens:
                    result.itens.append(
                        _ParsedItem(
                            codigo_origem=codigo,
                            descricao=descricao_clean,
                            unidade_medida=unidade_clean,
                            custo_base=custo,
                            tipo_recurso="SERVICO",
                        )
                    )
                    seen_itens.add(codigo)
            else:
                if current_parent_codigo is None:
                    result.avisos.append(
                        f"Linha {row_idx}: filho sem pai (ignorado): {codigo}"
                    )
                    continue

                if codigo not in seen_itens:
                    result.itens.append(
                        _ParsedItem(
                            codigo_origem=codigo,
                            descricao=descricao_clean,
                            unidade_medida=unidade_clean,
                            custo_base=custo,
                            tipo_recurso=tipo_recurso,
                        )
                    )
                    seen_itens.add(codigo)

                qty = float(coef) if coef is not None else 1.0
                result.relacoes.append(
                    _ParsedRelacao(
                        pai_codigo=current_parent_codigo,
                        filho_codigo=codigo,
                        quantidade_consumo=qty,
                        unidade_medida=unidade_clean,
                    )
                )

        wb.close()
        return self._store_and_build_response(result, "Composições TCPO - PINI.xlsx")

    # ── Parse: Converter em Data Center ───────────────────────────────────────

    def parse_converter_datacenter(self, file_bytes: bytes) -> EtlUploadResponse:
        """
        Parse auxiliary reference sheets from Converter em Data Center.xlsx.
        Each non-ENCARGOS sheet becomes a set of BaseTcpo items.
        Codes are prefixed (EPI-0001, EQP-0001, etc.) to avoid collisions.
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        result = _EtlParseResult()

        for sheet_name, (prefix, tipo, cc, dc, uc, pc) in _CONVERTER_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                result.avisos.append(f"Planilha '{sheet_name}' não encontrada — ignorada.")
                continue

            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_cod = row[cc]
                if raw_cod is None:
                    continue
                try:
                    codigo = f"{prefix}-{int(raw_cod):04d}"
                except (ValueError, TypeError):
                    continue

                descricao = str(row[dc]).strip() if row[dc] else ""
                unidade = str(row[uc]).strip() if (uc is not None and row[uc]) else "H" if tipo == "MO" else "UN"
                raw_price = row[pc] if pc < len(row) else None
                custo = float(raw_price) if raw_price is not None else 0.0

                result.itens.append(
                    _ParsedItem(
                        codigo_origem=codigo,
                        descricao=descricao,
                        unidade_medida=unidade,
                        custo_base=custo,
                        tipo_recurso=tipo,
                    )
                )

        wb.close()
        return self._store_and_build_response(result, "Converter em Data Center.xlsx")

    # ── Execute ────────────────────────────────────────────────────────────────

    async def execute_load(
        self, request: EtlExecuteRequest, db: AsyncSession
    ) -> EtlExecuteResponse:
        """
        Combines cached parse results and writes to referencia.* tables.

        UPSERT mode: safe incremental update; existing associations are never touched.
        REPLACE mode: clears BOM relationships + orphaned base_tcpo rows, then upserts.
        """
        t0 = time.monotonic()
        all_itens: list[_ParsedItem] = []
        all_relacoes: list[_ParsedRelacao] = []
        avisos: list[str] = []

        for token in filter(None, [request.parse_token_tcpo, request.parse_token_converter]):
            parsed = self._cache.pop(token, None)
            if parsed is None:
                raise ValueError(f"Parse token inválido ou expirado: {token}")
            all_itens.extend(parsed.itens)
            all_relacoes.extend(parsed.relacoes)
            avisos.extend(parsed.avisos)

        if not all_itens:
            raise ValueError("Nenhum item para carregar. Faça o upload primeiro.")

        # Deduplicate by codigo_origem (last occurrence wins for same code across files)
        dedup: dict[str, _ParsedItem] = {}
        for item in all_itens:
            dedup[item.codigo_origem] = item
        unique_itens = list(dedup.values())
        all_codes = [i.codigo_origem for i in unique_itens]

        # ── REPLACE: clear BOM and orphaned items ─────────────────────────────
        if request.mode == EtlMode.REPLACE:
            await db.execute(text("DELETE FROM referencia.composicao_base"))
            await db.execute(
                text(
                    """
                    DELETE FROM referencia.base_tcpo
                    WHERE id NOT IN (
                        SELECT DISTINCT item_referencia_id
                        FROM operacional.associacao_inteligente
                    )
                    """
                )
            )

        # ── Count existing to report inserts vs updates ───────────────────────
        existing_result = await db.execute(
            text(
                "SELECT codigo_origem FROM referencia.base_tcpo "
                "WHERE codigo_origem = ANY(:codes)"
            ),
            {"codes": all_codes},
        )
        existing_codes: set[str] = {row[0] for row in existing_result.fetchall()}
        inseridos = sum(1 for i in unique_itens if i.codigo_origem not in existing_codes)
        atualizados = len(unique_itens) - inseridos

        # ── Batch upsert base_tcpo ────────────────────────────────────────────
        for chunk_start in range(0, len(unique_itens), _BATCH_SIZE):
            chunk = unique_itens[chunk_start : chunk_start + _BATCH_SIZE]
            rows = [
                {
                    "id": uuid.uuid4(),
                    "codigo_origem": i.codigo_origem,
                    "descricao": i.descricao,
                    "unidade_medida": i.unidade_medida,
                    "custo_base": i.custo_base,
                    "tipo_recurso": i.tipo_recurso,
                }
                for i in chunk
            ]
            stmt = pg_insert(BaseTcpo).values(rows)
            stmt = stmt.on_conflict_do_update(
                index_elements=["codigo_origem"],
                set_={
                    "descricao": stmt.excluded.descricao,
                    "unidade_medida": stmt.excluded.unidade_medida,
                    "custo_base": stmt.excluded.custo_base,
                    "tipo_recurso": stmt.excluded.tipo_recurso,
                    "updated_at": func.now(),
                },
            )
            await db.execute(stmt)

        # ── Resolve codigo_origem → UUID for relationships ────────────────────
        relacoes_inseridas = 0
        if all_relacoes:
            all_rel_codes = list(
                {r.pai_codigo for r in all_relacoes}
                | {r.filho_codigo for r in all_relacoes}
            )
            mapping_result = await db.execute(
                select(BaseTcpo.id, BaseTcpo.codigo_origem).where(
                    BaseTcpo.codigo_origem.in_(all_rel_codes)
                )
            )
            code_to_id: dict[str, uuid.UUID] = {
                row[1]: row[0] for row in mapping_result.fetchall()
            }

            # UPSERT mode: delete existing BOM for parents we're about to reload
            if request.mode == EtlMode.UPSERT:
                parent_ids = list(
                    {
                        code_to_id[r.pai_codigo]
                        for r in all_relacoes
                        if r.pai_codigo in code_to_id
                    }
                )
                if parent_ids:
                    await db.execute(
                        text(
                            "DELETE FROM referencia.composicao_base "
                            "WHERE servico_pai_id = ANY(:ids)"
                        ),
                        {"ids": parent_ids},
                    )

            # Batch insert composicao_base
            valid_relacoes = []
            for rel in all_relacoes:
                pai_id = code_to_id.get(rel.pai_codigo)
                filho_id = code_to_id.get(rel.filho_codigo)
                if not pai_id or not filho_id:
                    avisos.append(
                        f"Relação ignorada — código ausente: "
                        f"{rel.pai_codigo} → {rel.filho_codigo}"
                    )
                    continue
                valid_relacoes.append(
                    {
                        "id": uuid.uuid4(),
                        "servico_pai_id": pai_id,
                        "insumo_filho_id": filho_id,
                        "quantidade_consumo": rel.quantidade_consumo,
                        "unidade_medida": rel.unidade_medida,
                    }
                )

            for chunk_start in range(0, len(valid_relacoes), _BATCH_SIZE):
                chunk = valid_relacoes[chunk_start : chunk_start + _BATCH_SIZE]
                await db.execute(pg_insert(ComposicaoBase).values(chunk))
                relacoes_inseridas += len(chunk)

        await db.commit()

        # ── Optional embedding recompute ──────────────────────────────────────
        embeddings_computados = 0
        if request.recomputar_embeddings:
            try:
                from app.services.servico_catalog_service import servico_catalog_service  # noqa: PLC0415

                embeddings_computados = await servico_catalog_service.compute_all_embeddings(db)
            except Exception as exc:  # noqa: BLE001
                logger.warning("etl.embeddings_failed", error=str(exc))
                avisos.append(f"Embeddings não computados: {exc}")

        duracao = time.monotonic() - t0
        logger.info(
            "etl.execute_complete",
            mode=request.mode,
            inseridos=inseridos,
            atualizados=atualizados,
            relacoes=relacoes_inseridas,
            embeddings=embeddings_computados,
            duracao=round(duracao, 2),
        )
        return EtlExecuteResponse(
            mode=request.mode,
            itens_inseridos=inseridos,
            itens_atualizados=atualizados,
            relacoes_inseridas=relacoes_inseridas,
            embeddings_computados=embeddings_computados,
            duracao_segundos=round(duracao, 2),
            avisos=avisos[:50],
        )

    # ── Status ─────────────────────────────────────────────────────────────────

    async def get_status(self, db: AsyncSession) -> EtlStatusResponse:
        result = await db.execute(
            text(
                """
                SELECT
                    (SELECT COUNT(*) FROM referencia.base_tcpo)::int,
                    (SELECT COUNT(*) FROM referencia.composicao_base)::int,
                    (SELECT COUNT(*) FROM referencia.tcpo_embeddings)::int,
                    (SELECT MAX(created_at) FROM referencia.base_tcpo)
                """
            )
        )
        row = result.fetchone()
        return EtlStatusResponse(
            total_itens_base_tcpo=row[0] or 0,
            total_composicoes_base=row[1] or 0,
            total_embeddings=row[2] or 0,
            ultima_carga=row[3],
        )

    # ── Internal helpers ───────────────────────────────────────────────────────

    def _store_and_build_response(
        self, result: _EtlParseResult, arquivo: str
    ) -> EtlUploadResponse:
        token = str(uuid.uuid4())
        self._cache[token] = result
        preview = EtlParsePreview(
            total_itens=len(result.itens),
            total_relacoes=len(result.relacoes),
            itens_amostra=[
                EtlItemPreview(
                    codigo_origem=i.codigo_origem,
                    descricao=i.descricao,
                    unidade_medida=i.unidade_medida,
                    custo_base=i.custo_base,
                    tipo_recurso=i.tipo_recurso,
                )
                for i in result.itens[:_SAMPLE_SIZE]
            ],
            relacoes_amostra=[
                EtlRelacaoPreview(
                    pai_codigo=r.pai_codigo,
                    filho_codigo=r.filho_codigo,
                    quantidade_consumo=r.quantidade_consumo,
                    unidade_medida=r.unidade_medida,
                )
                for r in result.relacoes[:_SAMPLE_SIZE]
            ],
            avisos=result.avisos[:20],
        )
        return EtlUploadResponse(arquivo=arquivo, parse_preview=preview, parse_token=token)


# Module-level singleton
etl_service = EtlService()
