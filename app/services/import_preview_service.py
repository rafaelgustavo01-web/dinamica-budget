import io
import re
from difflib import SequenceMatcher

import openpyxl

from app.schemas.admin import (
    FieldMappingPreview,
    ImportPreviewResponse,
    ImportSourceType,
    SheetPreview,
)


def _normalize(text: object) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().upper()


def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def _expected_headers_for(source_type: ImportSourceType) -> dict[str, list[str]]:
    if source_type == ImportSourceType.TCPO:
        return {
            "Composições sintéticas": ["CÓDIGOS", "DESCRIÇÃO RESUMO", "UNIDADE", "PREÇO"],
            "Composições analíticas": [
                "CÓDIGO",
                "DESCRIÇÃO",
                "CLASS",
                "UNIDADE",
                "COEF.",
                "PREÇO(R$)",
                "PREÇO TOTAL (R$)",
            ],
        }
    return {
        "MÃO DE OBRA": ["DESCRIÇÃO", "QUANTIDADE", "SALARIO", "CUSTO UNITARIO (H)", "CUSTO MENSAL"],
        "EQUIPAMENTOS": ["CÓDIGO", "EQUIPAMENTO", "CONSUMO", "ALUGUEL", "MÊS"],
        "ENCARGOS HORISTA": ["GRUPOS", "DISCRIMINAÇÃO DO ENCARGO", "TAXA (%)"],
        "ENCARGOS MENSALISTA": ["GRUPOS", "DISCRIMINAÇÃO DO ENCARGO", "TAXA (%)"],
        "EPI-UNIFORME": ["EPI", "UNID", "CUSTO UNITÁRIO", "QTDE", "CUSTO COM EPI(MÊS)"],
        "FERRAMENTAS": ["ITEM", "DESCRIÇÃO", "UNID.", "QUANT.", "PREÇO", "PREÇO TOTAL"],
        "MOBILIZAÇÃO": ["DESCRIÇÃO", "FUNÇÃO", "QUANTIDADE"],
    }


def _target_fields_for(source_type: ImportSourceType) -> list[str]:
    if source_type == ImportSourceType.TCPO:
        return [
            "codigo_origem",
            "descricao",
            "unidade_medida",
            "custo_unitario",
            "tipo_recurso",
            "quantidade_consumo",
        ]
    return [
        "descricao_funcao",
        "quantidade",
        "salario",
        "encargos_percent",
        "custo_unitario_h",
        "custo_mensal",
        "codigo",
        "equipamento",
        "consumo_l_h",
        "aluguel_r_h",
        "taxa_percent",
        "epi",
        "preco_total",
        "coluna_funcao",
    ]


_FIELD_HINTS: dict[str, list[str]] = {
    "codigo_origem": ["CÓDIGO", "CÓDIGOS"],
    "descricao": ["DESCRIÇÃO", "DESCRIÇÃO RESUMO"],
    "unidade_medida": ["UNIDADE", "UNID"],
    "custo_unitario": ["PREÇO", "PREÇO(R$)", "ALUGUEL"],
    "tipo_recurso": ["CLASS", "TIPO"],
    "quantidade_consumo": ["COEF", "QUANT"],
    "descricao_funcao": ["DESCRIÇÃO", "FUNÇÃO"],
    "quantidade": ["QUANTIDADE", "QTDE", "QUANT."],
    "salario": ["SALARIO"],
    "encargos_percent": ["ENCARGOS", "TAXA (%)"],
    "custo_unitario_h": ["CUSTO UNITARIO (H)", "CUSTO UNITÁRIO"],
    "custo_mensal": ["CUSTO MENSAL", "ALUGUEL MENSAL"],
    "codigo": ["CÓDIGO"],
    "equipamento": ["EQUIPAMENTO"],
    "consumo_l_h": ["CONSUMO"],
    "aluguel_r_h": ["ALUGUEL"],
    "taxa_percent": ["TAXA (%)"],
    "epi": ["EPI"],
    "preco_total": ["PREÇO TOTAL"],
    "coluna_funcao": ["ENG", "TEC", "ADM", "OFI", "AJUD"],
}


def _find_header_row(ws, expected_headers: list[str], max_scan: int = 25) -> tuple[int, list[str]]:
    expected = [_normalize(h) for h in expected_headers]
    best_row = 1
    best_score = -1.0
    best_headers: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row or 1), values_only=True), start=1):
        headers = [_normalize(v) for v in row if v not in (None, "")]
        if not headers:
            continue
        score = 0.0
        for exp in expected:
            score += max((_similarity(exp, h) for h in headers), default=0.0)
        if score > best_score:
            best_score = score
            best_row = i
            best_headers = headers

    return best_row, best_headers


def _map_headers(headers: list[str], target_fields: list[str]) -> list[FieldMappingPreview]:
    mapped: list[FieldMappingPreview] = []

    for h in headers:
        best_field = ""
        best_score = 0.0
        for field in target_fields:
            hints = _FIELD_HINTS.get(field, [field])
            score = max((_similarity(h, _normalize(x)) for x in hints), default=0.0)
            if score > best_score:
                best_score = score
                best_field = field
        if best_score >= 0.45:
            mapped.append(
                FieldMappingPreview(
                    source_header=h,
                    target_field=best_field,
                    confidence=round(min(best_score, 1.0), 4),
                )
            )

    return mapped


def generate_import_preview(source_type: ImportSourceType, file_name: str, file_bytes: bytes) -> ImportPreviewResponse:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    expected_map = _expected_headers_for(source_type)
    target_fields = _target_fields_for(source_type)

    sheets: list[SheetPreview] = []
    warnings: list[str] = []
    total_rows = 0
    estimated_records = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total_rows += ws.max_row or 0

        expected_headers = expected_map.get(sheet_name, [])
        header_row, header_values = _find_header_row(ws, expected_headers)
        mapped = _map_headers(header_values, target_fields)

        sample_rows: list[list[str]] = []
        for row in ws.iter_rows(
            min_row=min((header_row + 1), ws.max_row or 1),
            max_row=min((header_row + 4), ws.max_row or 1),
            values_only=True,
        ):
            values = [_normalize(v) for v in row if v not in (None, "")]
            if values:
                sample_rows.append(values[:12])

        estimated = max((ws.max_row or 0) - header_row, 0)
        estimated_records += estimated

        if expected_headers and len(mapped) < max(1, len(expected_headers) // 2):
            warnings.append(
                f"Aba '{sheet_name}' com mapeamento fraco ({len(mapped)} campos reconhecidos)."
            )

        sheets.append(
            SheetPreview(
                sheet_name=sheet_name,
                total_rows=ws.max_row or 0,
                header_row=header_row,
                estimated_records=estimated,
                mapped_fields=mapped,
                sample_rows=sample_rows,
            )
        )

    return ImportPreviewResponse(
        source_type=source_type,
        file_name=file_name,
        total_rows=total_rows,
        estimated_records=estimated_records,
        warnings=warnings,
        sheets=sheets,
    )
