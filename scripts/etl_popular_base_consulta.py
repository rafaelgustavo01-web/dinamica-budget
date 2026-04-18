import argparse
import os
import re
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import unquote

import openpyxl
import sqlalchemy as sa
from sqlalchemy.engine import make_url


def parse_env_file(path: Path) -> dict[str, str]:
    data: dict[str, str] = {}
    if not path.exists():
        return data
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        data[key.strip()] = value.strip()
    return data


def resolve_database_url(explicit: str | None) -> str:
    if explicit:
        return explicit

    env = os.environ.get("DATABASE_URL")
    if env:
        return env

    env_candidates = [
        Path(r"C:\DinamicaBudget\.env"),
        Path(r"C:\Dinamica-Budget\.env"),
        Path(r"C:\Dinamica-Budget\.env.example"),
    ]
    for candidate in env_candidates:
        values = parse_env_file(candidate)
        if values.get("DATABASE_URL"):
            return values["DATABASE_URL"]

    raise RuntimeError("DATABASE_URL nao encontrado em variavel de ambiente nem em .env")


def to_sync_url(database_url: str) -> str:
    parsed = make_url(database_url)
    pwd = parsed.password
    if pwd is not None:
        pwd = unquote(pwd)
    sync = parsed.set(drivername="postgresql+psycopg2", password=pwd)
    return sync.render_as_string(hide_password=False)


def norm_code(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", "", text)
    return text


def as_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("R$", "").replace("%", "").strip()
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def derive_tipo_recurso(classificacao: str, descricao: str) -> str:
    cls = (classificacao or "").upper()
    desc = (descricao or "").upper()

    if cls.startswith("SER"):
        return "SERVICO"
    if cls.startswith("EQ"):
        return "EQUIPAMENTO"
    if "FERR" in cls or "FERR" in desc:
        return "FERRAMENTA"
    if any(k in desc for k in ["PEDREIRO", "ENGENHEIRO", "AJUDANTE", "ELETRICISTA", "MONTADOR"]):
        return "MO"
    return "INSUMO"


def create_carga(con: sa.Connection, fonte_arquivo: str, tipo_fonte: str) -> uuid.UUID:
    carga_id = uuid.uuid4()
    con.execute(
        sa.text(
            """
            INSERT INTO etl_carga (id, fonte_arquivo, tipo_fonte, status, iniciado_em)
            VALUES (:id, :fonte_arquivo, :tipo_fonte, 'EM_PROCESSAMENTO', now())
            """
        ),
        {
            "id": carga_id,
            "fonte_arquivo": fonte_arquivo,
            "tipo_fonte": tipo_fonte,
        },
    )
    return carga_id


def close_carga(con: sa.Connection, carga_id: uuid.UUID, status: str, lidas: int, carregadas: int, mensagem: str | None) -> None:
    con.execute(
        sa.text(
            """
            UPDATE etl_carga
               SET status = :status,
                   finalizado_em = now(),
                   linhas_lidas = :lidas,
                   linhas_carregadas = :carregadas,
                   mensagem = :mensagem
             WHERE id = :id
            """
        ),
        {
            "id": carga_id,
            "status": status,
            "lidas": lidas,
            "carregadas": carregadas,
            "mensagem": mensagem,
        },
    )


def load_existing_servicos(con: sa.Connection) -> dict[str, uuid.UUID]:
    rows = con.execute(sa.text("SELECT codigo_origem, id FROM servico_tcpo")).mappings().all()
    return {row["codigo_origem"]: row["id"] for row in rows}


def upsert_servico_tcpo(
    con: sa.Connection,
    code_to_id: dict[str, uuid.UUID],
    codigo: str,
    descricao: str,
    unidade: str,
    custo: Decimal | None,
    tipo_recurso: str,
) -> uuid.UUID:
    if not codigo:
        raise ValueError("codigo_origem vazio")

    sid = code_to_id.get(codigo)
    if sid is None:
        sid = uuid.uuid4()
        con.execute(
            sa.text(
                """
                INSERT INTO servico_tcpo (
                    id, codigo_origem, descricao, unidade_medida, custo_unitario,
                    categoria_id, origem, status_homologacao, cliente_id,
                    tipo_recurso, descricao_tokens, created_at, updated_at
                ) VALUES (
                    :id, :codigo, :descricao, :unidade, :custo,
                    NULL, 'TCPO', 'APROVADO', NULL,
                    :tipo_recurso, NULL, now(), now()
                )
                """
            ),
            {
                "id": sid,
                "codigo": codigo,
                "descricao": descricao,
                "unidade": unidade or "UN",
                "custo": custo or Decimal("0"),
                "tipo_recurso": tipo_recurso,
            },
        )
        code_to_id[codigo] = sid
    else:
        con.execute(
            sa.text(
                """
                UPDATE servico_tcpo
                   SET descricao = :descricao,
                       unidade_medida = :unidade,
                       custo_unitario = :custo,
                       tipo_recurso = :tipo_recurso,
                       origem = 'TCPO',
                       status_homologacao = 'APROVADO',
                       cliente_id = NULL,
                       updated_at = now()
                 WHERE id = :id
                """
            ),
            {
                "id": sid,
                "descricao": descricao,
                "unidade": unidade or "UN",
                "custo": custo or Decimal("0"),
                "tipo_recurso": tipo_recurso,
            },
        )
    return sid


def parse_tcpo_workbook(path: Path) -> tuple[dict[str, dict], list[dict], int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    items: dict[str, dict] = {}
    edges: list[dict] = []
    lidas = 0

    ws_sint = wb["Composições sintéticas"]
    for row in ws_sint.iter_rows(min_row=2, values_only=True):
        lidas += 1
        codigo = norm_code(row[0])
        descricao = normalize_text(row[1])
        unidade = normalize_text(row[2])
        preco = as_decimal(row[3])
        if not codigo or not descricao:
            continue
        items[codigo] = {
            "codigo": codigo,
            "descricao": descricao,
            "unidade": unidade,
            "preco": preco,
            "tipo_recurso": "SERVICO",
        }

    ws_ana = wb["Composições analíticas"]
    current_parent: str | None = None
    code_pattern = re.compile(r"^[0-9A-Za-z]{8,}$")

    for row in ws_ana.iter_rows(min_row=2, values_only=True):
        lidas += 1
        codigo = norm_code(row[0])
        descricao = normalize_text(row[1])
        classificacao = normalize_text(row[2]).upper()
        unidade = normalize_text(row[3])
        coef = as_decimal(row[4])
        preco = as_decimal(row[5])

        if not codigo or not code_pattern.match(codigo):
            continue

        tipo_recurso = derive_tipo_recurso(classificacao, descricao)

        items.setdefault(
            codigo,
            {
                "codigo": codigo,
                "descricao": descricao,
                "unidade": unidade,
                "preco": preco,
                "tipo_recurso": tipo_recurso,
            },
        )

        if classificacao.startswith("SER"):
            current_parent = codigo
            continue

        if current_parent and current_parent != codigo and coef is not None:
            edges.append(
                {
                    "parent": current_parent,
                    "child": codigo,
                    "coef": coef,
                    "unidade": unidade or "UN",
                }
            )

    return items, edges, lidas


def ensure_versao_tcpo(con: sa.Connection, servico_id: uuid.UUID) -> uuid.UUID:
    row = con.execute(
        sa.text(
            """
            SELECT id
              FROM versao_composicao
             WHERE servico_id = :servico_id
               AND numero_versao = 1
               AND origem = 'TCPO'
               AND cliente_id IS NULL
             LIMIT 1
            """
        ),
        {"servico_id": servico_id},
    ).scalar_one_or_none()

    if row:
        return row

    vid = uuid.uuid4()
    con.execute(
        sa.text(
            """
            INSERT INTO versao_composicao (
                id, servico_id, numero_versao, origem, cliente_id, is_ativa, criado_por_id, criado_em
            ) VALUES (
                :id, :servico_id, 1, 'TCPO', NULL, TRUE, NULL, now()
            )
            """
        ),
        {"id": vid, "servico_id": servico_id},
    )
    return vid


def insert_composicao_if_missing(
    con: sa.Connection,
    servico_pai_id: uuid.UUID,
    insumo_filho_id: uuid.UUID,
    quantidade: Decimal,
    unidade: str,
    versao_id: uuid.UUID,
) -> bool:
    exists = con.execute(
        sa.text(
            """
            SELECT 1
              FROM composicao_tcpo
             WHERE versao_id = :versao_id
               AND servico_pai_id = :servico_pai_id
               AND insumo_filho_id = :insumo_filho_id
             LIMIT 1
            """
        ),
        {
            "versao_id": versao_id,
            "servico_pai_id": servico_pai_id,
            "insumo_filho_id": insumo_filho_id,
        },
    ).scalar_one_or_none()

    if exists:
        return False

    con.execute(
        sa.text(
            """
            INSERT INTO composicao_tcpo (
                id, servico_pai_id, insumo_filho_id, quantidade_consumo, versao_id, unidade_medida
            ) VALUES (
                :id, :servico_pai_id, :insumo_filho_id, :quantidade, :versao_id, :unidade
            )
            """
        ),
        {
            "id": uuid.uuid4(),
            "servico_pai_id": servico_pai_id,
            "insumo_filho_id": insumo_filho_id,
            "quantidade": quantidade,
            "versao_id": versao_id,
            "unidade": unidade,
        },
    )
    return True


def etl_tcpo(con: sa.Connection, file_path: Path) -> tuple[int, int]:
    items, edges, lidas = parse_tcpo_workbook(file_path)
    carregadas = 0

    code_to_id = load_existing_servicos(con)

    for item in items.values():
        upsert_servico_tcpo(
            con,
            code_to_id,
            item["codigo"],
            item["descricao"],
            item["unidade"],
            item["preco"],
            item["tipo_recurso"],
        )
        carregadas += 1

    for edge in edges:
        parent_id = code_to_id.get(edge["parent"])
        child_id = code_to_id.get(edge["child"])
        if not parent_id or not child_id:
            continue
        versao_id = ensure_versao_tcpo(con, parent_id)
        inserted = insert_composicao_if_missing(
            con,
            parent_id,
            child_id,
            edge["coef"],
            edge["unidade"],
            versao_id,
        )
        if inserted:
            carregadas += 1

    return lidas, carregadas


def find_header_row(ws, expected: list[str], max_scan: int = 25) -> int:
    target = {normalize_text(e).upper() for e in expected}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row or 1), values_only=True), start=1):
        values = {normalize_text(c).upper() for c in row if c not in (None, "")}
        if target.issubset(values):
            return i
    return 1


def build_header_map(row_values: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, value in enumerate(row_values):
        key = normalize_text(value).upper()
        if key:
            mapping[key] = idx
    return mapping


def get_cell(row: tuple[object, ...], idx: int | None) -> object:
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def etl_pc(con: sa.Connection, file_path: Path, carga_id: uuid.UUID) -> tuple[int, int]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    lidas = 0
    carregadas = 0

    pc_id = uuid.uuid4()
    con.execute(
        sa.text(
            """
            INSERT INTO pc_cabecalho (id, etl_carga_id, nome_arquivo, data_referencia, versao_layout, observacao, criado_em)
            VALUES (:id, :etl_carga_id, :nome_arquivo, NULL, NULL, :observacao, now())
            """
        ),
        {
            "id": pc_id,
            "etl_carga_id": carga_id,
            "nome_arquivo": file_path.name,
            "observacao": "Carga ETL automatica da planilha PC",
        },
    )
    carregadas += 1

    # MAO DE OBRA
    ws = wb["MÃO DE OBRA"]
    header_row = find_header_row(ws, ["DESCRIÇÃO", "QUANTIDADE", "SALARIO"])
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    hm = build_header_map(list(header_values))

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        lidas += 1
        descricao = normalize_text(get_cell(row, hm.get("DESCRIÇÃO")))
        if not descricao:
            continue
        con.execute(
            sa.text(
                """
                INSERT INTO pc_mao_obra_item (
                    id, pc_cabecalho_id, descricao_funcao, quantidade, salario, previsao_reajuste,
                    encargos_percent, custo_unitario_h, custo_mensal, mobilizacao
                ) VALUES (
                    :id, :pc_cabecalho_id, :descricao, :quantidade, :salario, :reajuste,
                    :encargos, :custo_h, :custo_mensal, :mobilizacao
                )
                """
            ),
            {
                "id": uuid.uuid4(),
                "pc_cabecalho_id": pc_id,
                "descricao": descricao,
                "quantidade": as_decimal(get_cell(row, hm.get("QUANTIDADE"))),
                "salario": as_decimal(get_cell(row, hm.get("SALARIO"))),
                "reajuste": as_decimal(get_cell(row, hm.get("PREVISÃO DE REAJUSTE"))),
                "encargos": as_decimal(get_cell(row, hm.get("ENCARGOS - %"))),
                "custo_h": as_decimal(get_cell(row, hm.get("CUSTO UNITARIO (H)"))),
                "custo_mensal": as_decimal(get_cell(row, hm.get("CUSTO MENSAL"))),
                "mobilizacao": as_decimal(get_cell(row, hm.get("MOBILIZAÇÃO"))),
            },
        )
        carregadas += 1

    # EQUIPAMENTOS
    ws = wb["EQUIPAMENTOS"]
    horas_mes = None
    preco_gasolina = None
    preco_diesel = None
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row or 10), values_only=True):
        lidas += 1
        row_text = [normalize_text(v).upper() for v in row if v not in (None, "")]
        if not row_text:
            continue
        line = " | ".join(row_text)
        if "HORAS/MÊS" in line or "HORAS/MES" in line:
            for c in row:
                d = as_decimal(c)
                if d is not None:
                    horas_mes = d
                    break
        if "GASOLINA" in line:
            for c in row:
                d = as_decimal(c)
                if d is not None:
                    preco_gasolina = d
                    break
        if "DIESEL" in line:
            for c in row:
                d = as_decimal(c)
                if d is not None:
                    preco_diesel = d
                    break

    con.execute(
        sa.text(
            """
            INSERT INTO pc_equipamento_premissa (id, pc_cabecalho_id, horas_mes, preco_gasolina_l, preco_diesel_l)
            VALUES (:id, :pc_cabecalho_id, :horas_mes, :gasolina, :diesel)
            """
        ),
        {
            "id": uuid.uuid4(),
            "pc_cabecalho_id": pc_id,
            "horas_mes": horas_mes,
            "gasolina": preco_gasolina,
            "diesel": preco_diesel,
        },
    )
    carregadas += 1

    header_row = find_header_row(ws, ["CÓDIGO", "EQUIPAMENTO", "ALUGUEL (R$/h)"])
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    hm = build_header_map(list(header_values))

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        lidas += 1
        equipamento = normalize_text(get_cell(row, hm.get("EQUIPAMENTO")))
        if not equipamento:
            continue
        con.execute(
            sa.text(
                """
                INSERT INTO pc_equipamento_item (
                    id, pc_cabecalho_id, codigo, equipamento, combustivel_utilizado,
                    consumo_l_h, aluguel_r_h, combustivel_r_h, mao_obra_r_h,
                    hora_produtiva, hora_improdutiva, mes, aluguel_mensal
                ) VALUES (
                    :id, :pc_cabecalho_id, :codigo, :equipamento, :combustivel,
                    :consumo, :aluguel_h, :combustivel_h, :mao_h,
                    :hp, :hi, :mes, :aluguel_mes
                )
                """
            ),
            {
                "id": uuid.uuid4(),
                "pc_cabecalho_id": pc_id,
                "codigo": normalize_text(get_cell(row, hm.get("CÓDIGO"))),
                "equipamento": equipamento,
                "combustivel": normalize_text(get_cell(row, hm.get("COMBUSTÍVEL UTILIZADO"))),
                "consumo": as_decimal(get_cell(row, hm.get("CONSUMO (L/H)"))),
                "aluguel_h": as_decimal(get_cell(row, hm.get("ALUGUEL (R$/H)"))),
                "combustivel_h": as_decimal(get_cell(row, hm.get("COMBUSTÍVEL (R$/H)"))),
                "mao_h": as_decimal(get_cell(row, hm.get("MÃO DE OBRA (R$/H)"))),
                "hp": as_decimal(get_cell(row, hm.get("HORA PRODUTIVA"))),
                "hi": as_decimal(get_cell(row, hm.get("HORA IMPRODUTIVA"))),
                "mes": as_decimal(get_cell(row, hm.get("MÊS"))),
                "aluguel_mes": as_decimal(get_cell(row, hm.get("ALUGUEL MENSAL"))),
            },
        )
        carregadas += 1

    # ENCARGOS HORISTA / MENSALISTA
    for sheet_name, tipo in [("ENCARGOS HORISTA", "HORISTA"), ("ENCARGOS MENSALISTA", "MENSALISTA")]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            lidas += 1
            vals = [normalize_text(v) for v in row if v not in (None, "")]
            if len(vals) < 3:
                continue
            if vals[0].upper() in {"GRUPOS", "ENCARGOS"}:
                continue
            taxa = as_decimal(vals[-1])
            if taxa is None:
                continue
            grupo = vals[0] if vals else None
            codigo_grupo = vals[1] if len(vals) > 1 else None
            discriminacao = vals[2] if len(vals) > 2 else vals[0]
            con.execute(
                sa.text(
                    """
                    INSERT INTO pc_encargo_item (
                        id, pc_cabecalho_id, tipo_encargo, grupo, codigo_grupo, discriminacao_encargo, taxa_percent
                    ) VALUES (
                        :id, :pc_cabecalho_id, :tipo, :grupo, :codigo_grupo, :discriminacao, :taxa
                    )
                    """
                ),
                {
                    "id": uuid.uuid4(),
                    "pc_cabecalho_id": pc_id,
                    "tipo": tipo,
                    "grupo": grupo,
                    "codigo_grupo": codigo_grupo,
                    "discriminacao": discriminacao,
                    "taxa": taxa,
                },
            )
            carregadas += 1

    # EPI
    ws = wb["EPI-UNIFORME"]
    header_row = find_header_row(ws, ["EPI", "UNID", "CUSTO UNITÁRIO", "QTDE", "CUSTO COM EPI(MÊS)"])
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    hm = build_header_map(list(header_values))
    fixed = {
        "EPI",
        "UNID",
        "CUSTO UNITÁRIO",
        "QTDE",
        "VIDA ÚTIL (*)(MÊS)",
        "CUSTO COM EPI(MÊS)",
    }

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        lidas += 1
        epi = normalize_text(get_cell(row, hm.get("EPI")))
        if not epi:
            continue
        epi_id = uuid.uuid4()
        con.execute(
            sa.text(
                """
                INSERT INTO pc_epi_item (
                    id, pc_cabecalho_id, epi, unidade, custo_unitario, quantidade, vida_util_meses, custo_epi_mes
                ) VALUES (
                    :id, :pc_cabecalho_id, :epi, :unidade, :custo_unitario, :quantidade, :vida_util, :custo_mes
                )
                """
            ),
            {
                "id": epi_id,
                "pc_cabecalho_id": pc_id,
                "epi": epi,
                "unidade": normalize_text(get_cell(row, hm.get("UNID"))),
                "custo_unitario": as_decimal(get_cell(row, hm.get("CUSTO UNITÁRIO"))),
                "quantidade": as_decimal(get_cell(row, hm.get("QTDE"))),
                "vida_util": as_decimal(get_cell(row, hm.get("VIDA ÚTIL (*)(MÊS)"))),
                "custo_mes": as_decimal(get_cell(row, hm.get("CUSTO COM EPI(MÊS)"))),
            },
        )
        carregadas += 1

        for idx, value in enumerate(header_values):
            fn = normalize_text(value)
            if not fn or fn.upper() in fixed:
                continue
            flag = normalize_text(get_cell(row, idx))
            if not flag:
                continue
            con.execute(
                sa.text(
                    """
                    INSERT INTO pc_epi_distribuicao_funcao (id, pc_epi_item_id, funcao, aplica_flag)
                    VALUES (:id, :pc_epi_item_id, :funcao, :aplica_flag)
                    """
                ),
                {
                    "id": uuid.uuid4(),
                    "pc_epi_item_id": epi_id,
                    "funcao": fn,
                    "aplica_flag": flag,
                },
            )
            carregadas += 1

    # FERRAMENTAS
    ws = wb["FERRAMENTAS"]
    header_row = find_header_row(ws, ["ITEM", "DESCRIÇÃO", "UNID.", "QUANT.", "PREÇO", "PREÇO TOTAL"])
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    hm = build_header_map(list(header_values))

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        lidas += 1
        descricao = normalize_text(get_cell(row, hm.get("DESCRIÇÃO")))
        if not descricao:
            continue
        con.execute(
            sa.text(
                """
                INSERT INTO pc_ferramenta_item (
                    id, pc_cabecalho_id, item, descricao, unidade, quantidade, preco, preco_total
                ) VALUES (
                    :id, :pc_cabecalho_id, :item, :descricao, :unidade, :quantidade, :preco, :preco_total
                )
                """
            ),
            {
                "id": uuid.uuid4(),
                "pc_cabecalho_id": pc_id,
                "item": normalize_text(get_cell(row, hm.get("ITEM"))),
                "descricao": descricao,
                "unidade": normalize_text(get_cell(row, hm.get("UNID."))),
                "quantidade": as_decimal(get_cell(row, hm.get("QUANT."))),
                "preco": as_decimal(get_cell(row, hm.get("PREÇO"))),
                "preco_total": as_decimal(get_cell(row, hm.get("PREÇO TOTAL"))),
            },
        )
        carregadas += 1

    # MOBILIZACAO (snapshot inicial de quantidades por funcao)
    ws = wb["MOBILIZAÇÃO"]
    func_headers: list[str] = []
    qty_values: list[Decimal | None] = []
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row or 20), values_only=True):
        lidas += 1
        vals = [normalize_text(v) for v in row if v not in (None, "")]
        if not vals:
            continue
        if vals[0].upper() == "QUANTIDADE":
            qty_values = [as_decimal(v) for v in row[1:10]]
        if vals and vals[0].upper() == "DESCRIÇÃO":
            func_headers = [normalize_text(v) for v in row[2:11]]

    mob_id = uuid.uuid4()
    con.execute(
        sa.text(
            """
            INSERT INTO pc_mobilizacao_item (id, pc_cabecalho_id, descricao, funcao, tipo_mao_obra)
            VALUES (:id, :pc_cabecalho_id, :descricao, :funcao, :tipo_mao_obra)
            """
        ),
        {
            "id": mob_id,
            "pc_cabecalho_id": pc_id,
            "descricao": "Quantidade base por funcao",
            "funcao": "MOBILIZACAO",
            "tipo_mao_obra": "MISTA",
        },
    )
    carregadas += 1

    for i, fn in enumerate(func_headers):
        if not fn:
            continue
        q = qty_values[i] if i < len(qty_values) else None
        con.execute(
            sa.text(
                """
                INSERT INTO pc_mobilizacao_quantidade_funcao (id, pc_mobilizacao_item_id, coluna_funcao, quantidade)
                VALUES (:id, :pc_mobilizacao_item_id, :coluna_funcao, :quantidade)
                """
            ),
            {
                "id": uuid.uuid4(),
                "pc_mobilizacao_item_id": mob_id,
                "coluna_funcao": fn,
                "quantidade": q,
            },
        )
        carregadas += 1

    return lidas, carregadas


def main() -> None:
    parser = argparse.ArgumentParser(description="ETL da base de consulta TCPO + PC")
    parser.add_argument(
        "--database-url",
        help="Sobrescreve DATABASE_URL de ambiente/.env",
    )
    parser.add_argument(
        "--tcpo-file",
        default=r"C:\Dinamica-Budget\tabelas\Composições TCPO - PINI.xlsx",
    )
    parser.add_argument(
        "--pc-file",
        default=r"C:\Dinamica-Budget\tabelas\PC tabelas.xlsx",
    )
    parser.add_argument("--only-tcpo", action="store_true")
    parser.add_argument("--only-pc", action="store_true")
    args = parser.parse_args()

    db_url = resolve_database_url(args.database_url)
    sync_url = to_sync_url(db_url)

    tcpo_file = Path(args.tcpo_file)
    pc_file = Path(args.pc_file)

    run_tcpo = not args.only_pc
    run_pc = not args.only_tcpo

    if run_tcpo and not tcpo_file.exists():
        raise FileNotFoundError(f"Arquivo TCPO nao encontrado: {tcpo_file}")
    if run_pc and not pc_file.exists():
        raise FileNotFoundError(f"Arquivo PC nao encontrado: {pc_file}")

    engine = sa.create_engine(sync_url)

    if run_tcpo:
        with engine.begin() as con:
            carga_tcpo = create_carga(con, str(tcpo_file), "TCPO")
        try:
            with engine.begin() as con:
                lidas, carregadas = etl_tcpo(con, tcpo_file)
            with engine.begin() as con:
                close_carga(con, carga_tcpo, "CONCLUIDO", lidas, carregadas, "Carga TCPO finalizada")
            print(f"[OK] TCPO: lidas={lidas} carregadas={carregadas}")
        except Exception as exc:
            with engine.begin() as con:
                close_carga(con, carga_tcpo, "ERRO", 0, 0, str(exc)[:1000])
            raise

    if run_pc:
        with engine.begin() as con:
            carga_pc = create_carga(con, str(pc_file), "PC")
        try:
            with engine.begin() as con:
                lidas, carregadas = etl_pc(con, pc_file, carga_pc)
            with engine.begin() as con:
                close_carga(con, carga_pc, "CONCLUIDO", lidas, carregadas, "Carga PC finalizada")
            print(f"[OK] PC: lidas={lidas} carregadas={carregadas}")
        except Exception as exc:
            with engine.begin() as con:
                close_carga(con, carga_pc, "ERRO", 0, 0, str(exc)[:1000])
            raise

    print("[OK] ETL concluido com sucesso")


if __name__ == "__main__":
    main()
